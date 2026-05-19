"""Generate the before/after Excel workbook (4 columns).

Output: docs/security/before_after_changes.xlsx
Columns: location | before code (single tenant) | after code (multi tenant) | explanation
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter


OUTPUT = "docs/security/before_after_changes.xlsx"


# Each entry: (location, before, after, explanation)
ROWS = [
    # ─────────── Phase 1 triage ───────────
    (
        "api/BL/blcontroller.py (line ~2602)",
        "elif self.object_name == 'test_trigger':\n"
        "    return create_app('Ic2di7G72HEviqQpWV','organization_')",
        "(branch deleted — also removed the orphaned `create_app` import)",
        "A hardcoded debug backdoor. ANY authenticated user could "
        "hit /api/test_trigger and create an app under a fixed "
        "hardcoded organization ID. SECURITY_AUDIT_REPORT graded "
        "this CRITICAL.",
    ),
    (
        "version2/settings.py (line ~119)",
        "SECRET_KEY = env(\"SECRET_KEY\", default='django-insecure-w+hl+it=...')",
        "try:\n"
        "    SECRET_KEY = env(\"SECRET_KEY\")\n"
        "except Exception:\n"
        "    raise ValueError(\"SECRET_KEY must be set...\")\n"
        "if len(SECRET_KEY) < 50:\n"
        "    raise ValueError(\"SECRET_KEY is too short...\")",
        "The fallback was the well-known Django default string. "
        "Shipping with it = session cookies, password reset links, "
        "CSRF tokens are forgeable by anyone who reads Django docs. "
        "Now the app refuses to boot without a real key.",
    ),
    (
        "api/APIs/dispatcher.py (lines ~72-93)",
        "request.user_ = user\n"
        "if request.user_ and not request.user_.get(\"is_active\", True):\n"
        "    raise PermissionError(\"User account is inactive.\")",
        "request.user_ = user\n"
        "if not request.user_:\n"
        "    raise PermissionError(\"Authentication required.\")\n"
        "if not request.user_.get(\"is_active\", False):\n"
        "    raise PermissionError(\"User account is inactive.\")",
        "Two security holes: (a) when request.user_ was None the "
        "condition short-circuited to False, request flowed through "
        "with NO user; (b) is_active default was True, missing key "
        "treated user as active. Both legs now fail-closed.",
    ),
    (
        "version2/settings.py (Security Settings block)",
        "SECURE_HSTS_SECONDS = env.int(\"SECURE_HSTS_SECONDS\", default=0)\n"
        "SECURE_SSL_REDIRECT = env.bool(\"SECURE_SSL_REDIRECT\", default=False)\n"
        "CSRF_COOKIE_SECURE = env.bool(\"CSRF_COOKIE_SECURE\", default=False)",
        "_IS_PROD = (ENVIRONMENT == \"production\")\n"
        "SECURE_HSTS_SECONDS = env.int(\"SECURE_HSTS_SECONDS\", "
        "default=31536000 if _IS_PROD else 0)\n"
        "SECURE_SSL_REDIRECT = env.bool(\"SECURE_SSL_REDIRECT\", default=_IS_PROD)\n"
        "CSRF_COOKIE_SECURE = env.bool(\"CSRF_COOKIE_SECURE\", default=_IS_PROD)\n"
        "if _IS_PROD:\n"
        "    SECURE_PROXY_SSL_HEADER = (\"HTTP_X_FORWARDED_PROTO\", \"https\")",
        "Defaults flipped to ON in production. Previously, forgetting "
        "any single env var shipped without HSTS, without SSL redirect, "
        "or with non-Secure cookies. Now production has these hardened "
        "by default; non-prod stays relaxed for local dev.",
    ),
    (
        "version2/settings.py (after JWT validation)",
        "(no SCHEMA_AUTHORITY_ENFORCE check at startup)",
        "_schema_authority_enforce = os.getenv(\"SCHEMA_AUTHORITY_ENFORCE\", \"1\")\n"
        "if ENVIRONMENT == \"production\" and _schema_authority_enforce != \"1\":\n"
        "    raise ValueError(\n"
        "        \"SCHEMA_AUTHORITY_ENFORCE must be '1' in production...\"\n"
        "    )",
        "SCHEMA_AUTHORITY_ENFORCE=0 is a soak-only escape hatch that "
        "downgrades cross-tenant kwarg violations to log-and-continue. "
        "Misconfiguring prod with =0 would let cross-tenant access slip "
        "through. App now refuses to boot in prod with the wrong value.",
    ),

    # ─────────── Foundation modules ───────────
    (
        "version2/settings.py (DATABASES block)",
        "DATABASES = {'default': {...standard fields only...}}",
        "DATABASES = {'default': {\n"
        "    ...standard fields...,\n"
        "    'ATOMIC_REQUESTS': True,\n"
        "    'CONN_MAX_AGE': 0,\n"
        "}}",
        "ATOMIC_REQUESTS wraps every view in a transaction so the "
        "middleware's SET LOCAL ROLE / SET LOCAL search_path bind to "
        "the request scope. CONN_MAX_AGE=0 forces a fresh connection "
        "per request so pool reuse doesn't leak stale tenant state.",
    ),
    (
        "api/security/tenant_schema_middleware.py "
        "(process_view body)",
        "with connection.cursor() as cur:\n"
        "    cur.execute(\"SET search_path TO %s, public\", [schema])\n"
        "request._tenant_search_path_set = True",
        "with connection.cursor() as cur:\n"
        "    cur.execute(\"SET LOCAL search_path TO %s, public\", [schema])\n"
        "    cur.execute(\"SET LOCAL app.current_org_id = %s\", [org_id])\n"
        "    cur.execute(\"SET LOCAL ROLE %s\", [role_name])\n"
        "request._tenant_search_path_set = True\n"
        "request._tenant_org_id_set = True\n"
        "request._tenant_role_set = True",
        "Three statements per request: (1) search_path routes queries "
        "to the tenant schema; (2) app.current_org_id is read by RLS "
        "policies; (3) SET LOCAL ROLE makes Postgres itself refuse "
        "cross-tenant access. Three independent defence layers.",
    ),
    (
        "api/security/tenant_schema_middleware.py "
        "(process_response)",
        "with connection.cursor() as cur:\n"
        "    cur.execute(\"SET search_path TO public\")",
        "with connection.cursor() as cur:\n"
        "    cur.execute(\"RESET ROLE\")\n"
        "    cur.execute(\"RESET app.current_org_id\")\n"
        "    cur.execute(\"SET search_path TO public\")",
        "On the way out, undo every per-request setting so the next "
        "request that picks up this pooled connection doesn't inherit "
        "our tenant role / org id / search_path. Belt-and-suspenders "
        "since ATOMIC_REQUESTS already bounds SET LOCAL.",
    ),
    (
        "api/tenant_models/_base.py (TenantModel base)",
        "class TenantModel(models.Model):\n"
        "    class Meta:\n"
        "        abstract = True\n"
        "        managed = False\n"
        "        app_label = \"api\"",
        "class TenantManager(models.Manager):\n"
        "    def for_tenant(self, ctx):\n"
        "        if not ctx or not ctx.schema:\n"
        "            raise TenantContextMissing(...)\n"
        "        # verify connection search_path matches ctx.schema\n"
        "        if ctx.schema not in actual_search_path:\n"
        "            raise TenantContextMismatch(...)\n"
        "        return self.get_queryset()\n"
        "\n"
        "class TenantModel(models.Model):\n"
        "    objects = TenantManager()\n"
        "    class Meta:\n"
        "        abstract = True\n"
        "        managed = False",
        "A defence-in-depth check at the ORM layer. Even if middleware "
        "didn't run (Celery task, mgmt command, raw helper), for_tenant "
        "asks Postgres SHOW search_path and refuses if it doesn't match "
        "the asserted context. Catches forgotten tenant pinning.",
    ),

    # ─────────── Encryption at rest ───────────
    (
        "api/security/encrypted_fields.py (NEW)",
        "(no encryption fields existed)",
        "class _EncryptedFieldMixin:\n"
        "    def from_db_value(self, value, ...):\n"
        "        return decrypt_token(value)\n"
        "    def get_prep_value(self, value):\n"
        "        if value.startswith(ENCRYPTED_PREFIX):\n"
        "            return value  # idempotent\n"
        "        return encrypt_token(value)\n"
        "\n"
        "class EncryptedCharField(_EncryptedFieldMixin, models.CharField):\n"
        "    ...\n"
        "class EncryptedTextField(_EncryptedFieldMixin, models.TextField):\n"
        "    ...",
        "Drop-in Django field types that encrypt on save and decrypt "
        "on read. Idempotent (re-saving doesn't double-encrypt). "
        "Legacy plaintext passthrough so existing rows keep working "
        "during migration. Used by every secret column below.",
    ),
    (
        "api/models.py (SessionLog)",
        "class SessionLog(models.Model):\n"
        "    access_token = models.CharField(max_length=500)\n"
        "    refresh_token = models.CharField(max_length=500)",
        "from api.security.encrypted_fields import EncryptedCharField\n"
        "\n"
        "class SessionLog(models.Model):\n"
        "    access_token = EncryptedCharField(max_length=1024)\n"
        "    refresh_token = EncryptedCharField(max_length=1024)",
        "Every session's JWT + refresh token stored encrypted at rest. "
        "Pre-fix: a DB dump exposed every active session's bearer "
        "token in plaintext. Post-fix: dump produces Fernet ciphertext.",
    ),
    (
        "api/models.py (SessionLog + UserLoginHistory)",
        "class SessionLog(models.Model):\n"
        "    user = models.ForeignKey(User, ...)\n"
        "    # no organization_id column",
        "class SessionLog(models.Model):\n"
        "    user = models.ForeignKey(User, ...)\n"
        "    organization_id = models.CharField(\n"
        "        max_length=64, null=True, blank=True, db_index=True)\n"
        "# same for UserLoginHistory",
        "Adds the column that Row-Level Security policies key on. "
        "Without it, RLS can't filter rows in these shared `public` "
        "tables to one tenant's view. Backfilled from users.organization_id "
        "via manage.py backfill_organization_id.",
    ),
    (
        "api/tenant_models/integration.py (TelephonyConfig)",
        "authtoken = models.CharField(max_length=512, null=True, blank=True)\n"
        "sid = models.CharField(max_length=512, null=True, blank=True)",
        "authtoken = EncryptedTextField(null=True, blank=True)\n"
        "sid = EncryptedTextField(null=True, blank=True)",
        "Twilio account_sid + auth_token now ciphertext at rest. A "
        "DB read no longer gives an attacker direct access to the "
        "customer's Twilio account.",
    ),
    (
        "api/tenant_models/integration.py "
        "(UserGmailToken / UserOutlookToken)",
        "access_token = models.TextField(null=True, blank=True)\n"
        "refresh_token = models.TextField(null=True, blank=True)",
        "access_token = EncryptedTextField(null=True, blank=True)\n"
        "refresh_token = EncryptedTextField(null=True, blank=True)",
        "Google OAuth + Microsoft OAuth tokens for connected user "
        "mailboxes now encrypted. Pre-fix, a DB dump exposed every "
        "customer's inbox-access tokens.",
    ),
    (
        "api/tenant_models/shared.py (LeadCapture)",
        "page_access_token = models.TextField(null=True, blank=True)",
        "page_access_token = EncryptedTextField(null=True, blank=True)",
        "Facebook page access token (used for Lead Ads webhook) now "
        "encrypted. Pre-fix: DB read → ability to read every "
        "customer's Facebook page traffic.",
    ),
    (
        "sf_integration/models.py (SalesforceSettings)",
        "password = models.CharField(max_length=255)\n"
        "client_secret = models.CharField(max_length=255)",
        "from api.security.encrypted_fields import EncryptedCharField\n"
        "\n"
        "password = EncryptedCharField(max_length=1024)\n"
        "client_secret = EncryptedCharField(max_length=1024)",
        "Salesforce admin credentials encrypted. Pre-fix, anyone with "
        "DB read could log in to your Salesforce org with the same "
        "credentials.",
    ),
    (
        "api/models.py (User.app_password)",
        "app_password = models.CharField(max_length=128, blank=True, null=True)",
        "app_password = EncryptedCharField(max_length=512, blank=True, null=True)",
        "Per-user SMTP app password encrypted. Pre-fix: every user's "
        "outbound SMTP credentials in plaintext.",
    ),

    # ─────────── Database-enforced isolation (Phase 4) ───────────
    (
        "scripts/per_tenant_ddl/provision_tenant_role.sql (NEW)",
        "(one application role with full access to every tenant schema)",
        "CREATE ROLE tenant_${SCHEMA}_role NOLOGIN INHERIT;\n"
        "GRANT USAGE ON SCHEMA \"${SCHEMA}\" TO tenant_${SCHEMA}_role;\n"
        "GRANT SELECT, INSERT, UPDATE, DELETE\n"
        "    ON ALL TABLES IN SCHEMA \"${SCHEMA}\"\n"
        "    TO tenant_${SCHEMA}_role;\n"
        "GRANT SELECT, INSERT, UPDATE ON public.users\n"
        "    TO tenant_${SCHEMA}_role;\n"
        "-- narrow shared-table grants...\n"
        "GRANT tenant_${SCHEMA}_role TO bussus_app;",
        "Per-tenant Postgres role that has permissions ONLY on its "
        "own schema + a narrow whitelist of shared tables. The "
        "application's main role (bussus_app) is GRANTed membership "
        "so it can SET ROLE into any tenant role. After the middleware "
        "issues SET LOCAL ROLE, Postgres itself refuses cross-tenant "
        "queries with 'permission denied'.",
    ),
    (
        "api/migrations/0015_enable_rls_shared_tables.py (NEW)",
        "(shared public tables had no RLS; tenant role with table "
        "grant could read every tenant's rows)",
        "ALTER TABLE public.users ENABLE ROW LEVEL SECURITY;\n"
        "CREATE POLICY tenant_isolation\n"
        "    ON public.users\n"
        "    FOR ALL\n"
        "    USING  (organization_id = current_setting('app.current_org_id'))\n"
        "    WITH CHECK (organization_id = current_setting('app.current_org_id'));\n"
        "-- same for organizations, user_login_history, session_log, lead_capture",
        "Row-Level Security policies on the 5 shared `public` tables. "
        "Postgres silently adds 'WHERE organization_id = <current "
        "tenant>' to every SELECT/UPDATE/DELETE. WITH CHECK blocks "
        "writes that try to place a row under another tenant's org. "
        "Cross-tenant queries return 0 rows instead of leaking.",
    ),
    (
        "api/management/commands/provision_tenant_role.py (NEW)",
        "(no automation; tenant roles didn't exist)",
        "class Command(BaseCommand):\n"
        "    def handle(self, *args, **opts):\n"
        "        for schema in active_orgs:\n"
        "            sql.SQL('CREATE ROLE %I NOLOGIN INHERIT').format(...)\n"
        "            sql.SQL('GRANT USAGE ON SCHEMA {sch} TO {role}').format(...)\n"
        "            # ...all grants from provision_tenant_role.sql\n"
        "            sql.SQL('GRANT {role} TO {app}').format(...)",
        "Safe Python wrapper that substitutes the schema name via "
        "psycopg2.sql.Identifier (no string-format SQL injection). "
        "Supports --all (every active org) + --dry-run + --revoke. "
        "Idempotent — re-runs are safe. Run during tenant onboarding.",
    ),

    # ─────────── Auth gate + rate limit (Phase 2 + 8.A7) ───────────
    (
        "version2/settings.py (REST_FRAMEWORK block)",
        "REST_FRAMEWORK = {\n"
        "    'DEFAULT_AUTHENTICATION_CLASSES': (),\n"
        "    'DEFAULT_PERMISSION_CLASSES': (),\n"
        "}",
        "STRICT_AUTH = env.bool(\"STRICT_AUTH\", default=False)\n"
        "if STRICT_AUTH:\n"
        "    REST_FRAMEWORK = {\n"
        "        'DEFAULT_AUTHENTICATION_CLASSES': (\n"
        "            'authentication.custom_jwt_auth.CustomJWTAuthentication',\n"
        "        ),\n"
        "        'DEFAULT_PERMISSION_CLASSES': (\n"
        "            'rest_framework.permissions.IsAuthenticated',\n"
        "        ),\n"
        "    }",
        "Feature-flagged DRF global auth. With STRICT_AUTH=1, every "
        "view requires JWT by default; opt-out is explicit via "
        "permission_classes=[AllowAny]. Pre-fix, a new view shipped "
        "without permission_classes was wide open. Default OFF "
        "during rollout; operator flips to 1 after staging soak.",
    ),
    (
        "public/auth/login.py (LoginView decorator)",
        "@method_decorator(ratelimit(key='ip', rate='5/m', method='POST'), "
        "name='dispatch')\nclass LoginView(View):\n"
        "    def post(self, request, ...):\n"
        "        # no lockout check",
        "@method_decorator(\n"
        "    ratelimit(key='ip', rate='20/h', method='POST', block=True),\n"
        "    name='dispatch',\n)\n"
        "@method_decorator(\n"
        "    ratelimit(key='post:username', rate='5/h', method='POST', block=True),\n"
        "    name='dispatch',\n)\n"
        "class LoginView(View):\n"
        "    def post(self, request, ...):\n"
        "        if is_locked_out(username):\n"
        "            return JsonResponse(lockout_response_payload(username), status=429)",
        "Pre-fix, the @ratelimit decorator was a no-op — without "
        "block=True it just set request.limited and the view never "
        "checked. Now it returns 429 itself. Stacked IP + per-username "
        "keys catch both burst and credential stuffing. Progressive "
        "lockout via UserLoginHistory blocks even cross-IP attempts.",
    ),
    (
        "public/auth/lockout.py (NEW)",
        "(no lockout — unlimited failed attempts)",
        "def is_locked_out(email):\n"
        "    recent_failures = UserLoginHistory.objects.filter(\n"
        "        user__email__iexact=email,\n"
        "        login_time__gte=now() - timedelta(minutes=15),\n"
        "        login_type='failed',\n"
        "    ).count()\n"
        "    return recent_failures >= 5",
        "After 5 failed logins for the same email within 15 minutes, "
        "the account is locked for 15 minutes. Catches credential "
        "stuffing even when the attacker rotates IPs. Fail-OPEN on "
        "query error so a helper bug doesn't take down login.",
    ),

    # ─────────── SQL injection ───────────
    (
        "api/BL/computed_fields.py (push-down query builder)",
        "full_sql = (\n"
        "    f\"SELECT id FROM \\\"{parent_table}\\\" \"\n"
        "    f\"WHERE ({expr}) {op_sql} %s\"\n"
        ")\n"
        "cursor.execute(full_sql, [value])",
        "from api.BL.computed_fields_columns import assert_column, assert_operator\n"
        "from psycopg2 import sql\n"
        "\n"
        "op_spec = assert_operator(op_sql)  # whitelist check\n"
        "for ident in identifiers:\n"
        "    assert_column(schema, parent_table, ident)\n"
        "full_sql = sql.SQL(\"SELECT id FROM {tbl} WHERE ({expr}) \").format(\n"
        "    tbl=sql.Identifier(parent_table),\n"
        "    expr=sql.SQL(expr),\n"
        ") + sql.SQL(op_spec['sql']) + sql.SQL(\" %s\")",
        "Audit graded the f-string SQL CRITICAL — whitelist permitted "
        "parens/commas/operators, enough to inject UNION SELECT. Now "
        "every identifier passes through psycopg2.sql.Identifier; "
        "every column is verified against information_schema; the "
        "operator is on a hardcoded whitelist.",
    ),
    (
        "api/BL/computed_fields_columns.py (NEW)",
        "(no central allow-list; per-call ad-hoc validation)",
        "def get_allowed_columns(schema, table):\n"
        "    # query information_schema.columns, cache by (schema, table)\n"
        "    ...\n"
        "def assert_column(schema, table, column):\n"
        "    if column not in get_allowed_columns(schema, table):\n"
        "        raise InvalidIdentifierError(...)\n"
        "def assert_operator(op):\n"
        "    if op not in ALLOWED_COMPARISON_OPERATORS:\n"
        "        raise InvalidOperatorError(...)",
        "Authoritative allow-list pulled directly from "
        "information_schema (not a hand-maintained list that could "
        "drift). Cached 5min with DDL-invalidation hook. Fail-CLOSED "
        "on DB error — returns empty set so callers reject every "
        "identifier rather than risk letting one through.",
    ),

    # ─────────── Mass assignment ───────────
    (
        "api/BL/blcontroller.py (task POST)",
        "modified_data = {\n"
        "    **create_data,                  # ← user payload spread\n"
        "    'created_by_id': user_id,\n"
        "    'created_date': now,\n"
        "}\n"
        "post_data_sql('task', modified_data, ...)",
        "from api.BL.allowed_fields import sanitize_create_payload\n"
        "\n"
        "modified_data, _dropped = sanitize_create_payload(\n"
        "    create_data,\n"
        "    schema=schema_for_sanitize,\n"
        "    object_name='task',\n"
        "    user_id=user_id,\n"
        "    now=now,\n"
        ")\n"
        "post_data_sql('task', modified_data, ...)",
        "Pre-fix, {**user_payload} spread let an attacker set id, "
        "owner_id, organization_id, is_deleted, is_staff in their "
        "POST body. Now drops every key not on the per-tenant allow-"
        "list + drops every system-field denylist member + layers "
        "trusted system fields on top.",
    ),
    (
        "api/BL/allowed_fields.py (NEW)",
        "(no central sanitiser; each call site built create_data ad-hoc)",
        "SYSTEM_FIELDS_DENYLIST = frozenset({\n"
        "    'id', 'created_by_id', 'organization_id', 'is_deleted',\n"
        "    'owner_id', 'tenant_id', ...\n"
        "})\n"
        "def sanitize_create_payload(payload, schema, object_name, user_id):\n"
        "    allowed = get_allowed_create_fields(schema, object_name)\n"
        "    safe = {k: v for k, v in payload.items()\n"
        "           if k in allowed and k not in SYSTEM_FIELDS_DENYLIST}\n"
        "    safe['created_by_id'] = user_id  # platform value wins\n"
        "    safe['created_date'] = now\n"
        "    return safe, dropped_keys",
        "Two-layer defence. Allow-list comes from the per-tenant "
        "`fields` metadata table (admin-configured). Denylist is "
        "hardcoded — even if a tenant misconfigures and flags `id` "
        "as modifiable, the platform refuses. System fields layered "
        "AFTER the allow-list check so they always win.",
    ),

    # ─────────── File upload ───────────
    (
        "utils/file_handling.py (handle_file_upload)",
        "def handle_file_upload(file, **kwargs):\n"
        "    org_name = kwargs.get('org', {}).get('name', 'public')\n"
        "    upload_folder = os.path.join('uploads', org_name)\n"
        "    file_path = os.path.join(upload_folder, file.name)\n"
        "    # no size cap, no MIME check, raw user filename",
        "from utils.file_validation import validate_upload\n"
        "\n"
        "def handle_file_upload(file, **kwargs):\n"
        "    mime, safe_name = validate_upload(\n"
        "        file, kind=kwargs.get('kind'),\n"
        "        max_bytes=kwargs.get('max_bytes') or 50*1024*1024,\n"
        "    )\n"
        "    # UUID-prefixed path so safe_name can't collide\n"
        "    upload_folder = os.path.join('uploads',\n"
        "        sanitize(org_name), uuid.uuid4().hex)\n"
        "    file_path = os.path.join(upload_folder, safe_name)",
        "Every upload now validated: size cap (50MB default), MIME "
        "via python-magic (content sniffing, not Content-Type), "
        "filename sanitised (no path traversal), UUID-prefixed path "
        "(no collisions). Pre-fix: any auth'd user could POST a 10GB "
        "file with name '../../etc/passwd'.",
    ),
    (
        "utils/file_validation.py (NEW)",
        "(no central validator)",
        "MAX_BYTES = 10 * 1024 * 1024\n"
        "ALLOWED_MIME = {\n"
        "    'image': {'image/jpeg', 'image/png', ...},\n"
        "    'doc':   {'application/pdf', ...},\n"
        "}\n"
        "def validate_upload(file, kind='image', max_bytes=MAX_BYTES):\n"
        "    if file.size > max_bytes: raise ValidationError(...)\n"
        "    mime = magic.from_buffer(file.read(2048), mime=True)\n"
        "    if mime not in ALLOWED_MIME[kind]: raise ValidationError(...)\n"
        "    safe_name = sanitize_filename(file.name)\n"
        "    return mime, safe_name",
        "Centralised validation: size, MIME (content-sniffed via "
        "python-magic, not header-sniffed), filename sanitisation "
        "(strips path components + shell metacharacters). Falls back "
        "to extension check if libmagic isn't installed (warns).",
    ),
    (
        "Dockerfile",
        "FROM python:3.10-slim\nWORKDIR /app\nCOPY . .",
        "FROM python:3.10-slim\n"
        "WORKDIR /app\n"
        "RUN apt-get update \\\n"
        " && apt-get install -y --no-install-recommends libmagic1 \\\n"
        " && rm -rf /var/lib/apt/lists/*\n"
        "COPY . .",
        "libmagic1 is the system library python-magic depends on for "
        "content-based MIME detection. Without it, file validation "
        "falls back to extension-only checks (spoof-able).",
    ),

    # ─────────── Background work (Phase 6) ───────────
    (
        "api/celery_tasks/base.py (NEW)",
        "(bare @shared_task; no tenant guarantees)",
        "class TenantRequiredTask(Task):\n"
        "    def __call__(self, *args, **kwargs):\n"
        "        raw_ctx = kwargs.pop('_tenant_ctx', None)\n"
        "        if not raw_ctx:\n"
        "            raise RuntimeError(\n"
        "                f'Task {self.name!r} requires _tenant_ctx kwarg.')\n"
        "        ctx = TenantContext(**raw_ctx)\n"
        "        with with_tenant_schema(ctx.schema):\n"
        "            return self.run(ctx, *args, **kwargs)",
        "Celery base class that REFUSES to run without _tenant_ctx "
        "in kwargs. Wraps the task body in with_tenant_schema() so "
        "DB queries auto-scope to the right schema. Forgetting the "
        "kwarg becomes a hard runtime error instead of silent wrong-"
        "schema queries.",
    ),
    (
        "adminuser/tasks.py (log_user_login_async)",
        "@shared_task(bind=True, max_retries=3)\n"
        "def log_user_login_async(self, user_id, ...):\n"
        "    SessionLog.objects.create(\n"
        "        user=user, profile_id=..., ...)",
        "from api.celery_tasks.base import TenantRequiredTask\n"
        "\n"
        "@shared_task(base=TenantRequiredTask, bind=True, max_retries=3)\n"
        "def log_user_login_async(self, ctx, user_id, ...):\n"
        "    SessionLog.objects.create(\n"
        "        user=user,\n"
        "        organization_id=ctx.org_id,  # explicit\n"
        "        profile_id=..., ...)",
        "Writes to public.session_log + public.user_login_history "
        "(RLS-scoped tables). Without ctx.org_id explicitly set, RLS "
        "WITH CHECK would refuse the INSERT under FORCE RLS. Now "
        "the task carries tenant context end-to-end.",
    ),
    (
        "adminuser/LoginView.py (apply_async call)",
        "log_user_login_async.delay(\n"
        "    user_id=user_id, profile_id=..., ...)",
        "from api.celery_tasks.base import serialize_ctx\n"
        "from api.security.schema_authority import TenantContext\n"
        "\n"
        "# Look up schema from organizations\n"
        "ctx = TenantContext(\n"
        "    org_id=str(organization_id),\n"
        "    schema=schema_for_ctx,\n"
        "    profile_id=str(profile_id))\n"
        "log_user_login_async.apply_async(kwargs={\n"
        "    '_tenant_ctx': serialize_ctx(ctx),\n"
        "    'user_id': user_id, ...})",
        "Caller now builds a TenantContext from the user's org and "
        "passes it via _tenant_ctx. Without this, TenantRequiredTask "
        "would refuse to run the task at all.",
    ),

    # ─────────── Cache + storage namespacing (Phase 7) ───────────
    (
        "CacheService/tenant_cache.py (NEW)",
        "from django.core.cache import cache\n"
        "cache.set('user_perms', perms_dict, timeout=300)\n"
        "cache.get('user_perms')\n"
        "# no tenant prefix; two tenants could collide",
        "def _build_key(org_id, key):\n"
        "    return f'tenant:{org_id}:{key}'\n"
        "def tenant_set(ctx, key, value, timeout=None):\n"
        "    if not ctx or not ctx.org_id: raise TenantContextMissing(...)\n"
        "    cache.set(_build_key(ctx.org_id, key), value, timeout)\n"
        "def tenant_get(ctx, key, default=None):\n"
        "    return cache.get(_build_key(ctx.org_id, key), default)",
        "Every Redis key prefixed `tenant:<org_id>:<key>`. Pre-fix, "
        "tenant A caching by `user_perms` and tenant B caching by "
        "`user_perms` collided. Now `tenant:acme:user_perms` vs "
        "`tenant:beta:user_perms` — no collision.",
    ),

    # ─────────── Observability (10-risk) ───────────
    (
        "api/security/sentry_tags.py (NEW)",
        "(Sentry events had no tenant attribution)",
        "class SentryTenantTagMiddleware(MiddlewareMixin):\n"
        "    def process_request(self, request):\n"
        "        with sentry_sdk.configure_scope() as scope:\n"
        "            if hasattr(request, 'tenant_org_id'):\n"
        "                scope.set_tag('tenant_id', request.tenant_org_id)\n"
        "            if hasattr(request, 'tenant_schema'):\n"
        "                scope.set_tag('tenant_schema', request.tenant_schema)\n"
        "            user_ = getattr(request, 'user_', None)\n"
        "            if isinstance(user_, dict) and user_.get('id'):\n"
        "                scope.set_user({'id': user_['id']})",
        "Every Sentry error now tagged with tenant_id + tenant_schema + "
        "user_id. Pre-fix: all multi-tenant errors landed in one bucket "
        "and triage required reverse-engineering trace_id to find which "
        "tenant. Now Sentry filters per tenant out of the box.",
    ),
    (
        "api/security/cross_tenant_heartbeat.py (NEW)",
        "(no continuous Phase 4 verification)",
        "@shared_task(base=AdminTask)\n"
        "def run_cross_tenant_probe():\n"
        "    a, b = pick_two_active_orgs()\n"
        "    with connection.cursor() as cur:\n"
        "        cur.execute('SET LOCAL ROLE %s', [f'tenant_{a}_role'])\n"
        "        cur.execute('SET LOCAL app.current_org_id = %s', [a])\n"
        "        cur.execute('SELECT count(*) FROM public.users '\n"
        "                    'WHERE organization_id = %s', [b])\n"
        "        if cur.fetchone()[0] > 0:\n"
        "            sentry_sdk.capture_message(..., level='fatal')",
        "Hourly Celery task that PROACTIVELY tries to do a cross-tenant "
        "query as a tenant role. If RLS or the role grants ever break "
        "in prod, this task pages on-call within an hour. Continuous-"
        "verification companion to the unit tests.",
    ),

    # ─────────── Phase 5 adoption ───────────
    (
        "api/permissions/permissions.py "
        "(_get_object_details_orm, etc.)",
        "qs = PlatformObject.objects.filter(name=table_name)\n"
        "row = qs.values_list('id', 'label').first()\n"
        "return row",
        "from api.security.schema_authority import TenantContext\n"
        "\n"
        "ctx = TenantContext(org_id=schema, schema=schema, profile_id=None)\n"
        "qs = PlatformObject.objects.for_tenant(ctx).filter(name=table_name)\n"
        "row = qs.values_list('id', 'label').first()\n"
        "return row",
        "Pre-fix: naked .objects.filter() trusted that the middleware "
        "had set search_path. Post-fix: for_tenant(ctx) re-verifies "
        "search_path matches ctx.schema before returning the queryset. "
        "If middleware was bypassed, query is refused at TenantManager.",
    ),

    # ─────────── Handler registry (god-file split wave 1) ───────────
    (
        "api/BL/handlers/_base.py (NEW)",
        "(no per-domain handlers — 5,088-line god-file blcontroller.py)",
        "class DomainHandler:\n"
        "    OBJECT_NAMES: tuple[str, ...] = ()\n"
        "    def __init__(self, request, object_name, ctx=None):\n"
        "        self.request = request\n"
        "        self.object_name = object_name\n"
        "        self.ctx = ctx  # tenant context\n"
        "    def get(self, **kwargs): return NotImplementedForVerb\n"
        "    def post(self, data, **kwargs): return NotImplementedForVerb\n"
        "    # patch, delete...\n"
        "\n"
        "HANDLER_REGISTRY: dict[str, type[DomainHandler]] = {}",
        "Pattern for splitting the god-file. Each domain (Task, File, "
        "RecycleBin, etc.) becomes a subclass of DomainHandler with "
        "its OBJECT_NAMES tuple. BusinessLogicHandler dispatches "
        "through the registry. Wave 1 extracted Task; future waves "
        "extract more. Less coupling, smaller blast radius per bug.",
    ),

    # ─────────── URL auth gate scaffolding ───────────
    (
        "tests/security/public_urls.txt (NEW)",
        "(no enumerated public-URL list; relied on per-view discipline)",
        "# Health probes\n"
        "^/healthz$\n^/livez$\n^/readyz$\n"
        "# Auth flows\n"
        "^/v2/login$\n^/v2/auth/signup$\n^/v2/reset_password$\n"
        "# OTP\n^/v2/start$\n^/v2/verify$\n^/v2/resend$\n"
        "# OAuth callbacks\n"
        "^/v2/api/gmail/oauth/callback/$\n"
        "# Webhooks (HMAC-verified)\n^/v2/telephony/route$\n^/whatsapp/webhook/$",
        "Source of truth for legitimately-public URLs. Regression test "
        "(test_auth_required.py) walks the URL conf and flags any view "
        "that's AllowAny but not on the allowlist. Adding a new public "
        "URL forces explicit security review on the PR.",
    ),
]


def _border():
    side = Side(border_style="thin", color="CCCCCC")
    return Border(top=side, bottom=side, left=side, right=side)


def _header_fill():
    return PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")


def _alt_fill():
    return PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "before_after"

    headers = ["location", "before code (single tenant)",
               "after code (multi tenant)", "explanation"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    body_font = Font(name="Consolas", size=10)
    expl_font = Font(name="Calibri", size=10)
    border = _border()
    header_fill = _header_fill()
    alt_fill = _alt_fill()

    # Style header row
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="left", vertical="center", wrap_text=True,
        )
        cell.border = border
    ws.row_dimensions[1].height = 26

    # Body rows
    for i, (location, before, after, explanation) in enumerate(ROWS, start=2):
        ws.cell(row=i, column=1, value=location)
        ws.cell(row=i, column=2, value=before)
        ws.cell(row=i, column=3, value=after)
        ws.cell(row=i, column=4, value=explanation)

        for col in range(1, 5):
            cell = ws.cell(row=i, column=col)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True,
            )
            cell.border = border
            cell.font = (body_font if col in (2, 3) else expl_font)
            if i % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    widths = {"A": 42, "B": 60, "C": 60, "D": 56}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Freeze the header
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"Generated: {OUTPUT}  ({len(ROWS)} rows)")


if __name__ == "__main__":
    build()
