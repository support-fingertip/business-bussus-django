"""
Build an Excel audit workbook of every entry in the `object` registry.

Reads the seed data in sqlfiles/objects.sql, cross-references the Django
models actually defined in the codebase, and emits an .xlsx with four
sheets: Summary, All_Objects, Migration_Plan, Registry_Gaps.
"""
import os
import re
import sys
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = "/home/user/business-bussus-django"
OBJECTS_SQL = os.path.join(ROOT, "sqlfiles", "objects.sql")
OUT_XLSX = os.path.join(ROOT, "objects_audit.xlsx")


# Tables already defined as Django models in the codebase (db_table -> source path)
DJANGO_MODELED = {
    "organizations":         "api/models.py:Organization",
    "users":                 "api/models.py:User",
    "session_log":           "api/models.py:SessionLog",
    "user_login_history":    "api/models.py:UserLoginHistory",
    "facebook_lead":         "facebook/models.py",
    "facebookleadwebhooks":  "facebook/models.py",
    "message":               "whatsapp/models.py",
    "channel":               "whatsapp/models.py",
    "webhook":               "whatsapp/models.py",
    "salesforce_metadata":   "sf_integration/models.py",
    "salesforce_settings":   "sf_integration/models.py",
    "salesforce_sync":       "sf_integration/models.py",
}


# Functional grouping for setup tables (used for the Subcategory column)
SETUP_GROUPS = {
    # Platform metadata
    "object":                "Platform Metadata",
    "fields":                "Platform Metadata",
    "tables_metadata":       "Platform Metadata",
    "columns_metadata":      "Platform Metadata",
    "setup":                 "Platform Metadata",
    "custom_metadata":       "Platform Metadata",
    "custom_setting":        "Platform Metadata",
    # Authorization & sharing
    "profile":               "Authorization",
    "permission_sets":       "Authorization",
    "roles":                 "Authorization",
    "user_group":            "Authorization",
    "user_group_users":      "Authorization",
    "users_user_permissions":"Authorization",
    "object_permissions":    "Authorization",
    "field_permissions":     "Authorization",
    "tab_permissions":       "Authorization",
    "app_permissions":       "Authorization",
    "sharing_records":       "Authorization",
    "sharing_rules":         "Authorization",
    "owd":                   "Authorization",
    "auth_group":            "Authorization (Django)",
    "auth_group_permissions":"Authorization (Django)",
    "auth_permission":       "Authorization (Django)",
    # Identity
    "users":                 "Identity",
    # UI / layout / navigation
    "app":                   "UI / Layout",
    "tabs":                  "UI / Layout",
    "page_layouts":          "UI / Layout",
    "page_builder_assignment":"UI / Layout",
    "layout_assignment":     "UI / Layout",
    "lightning_pages":       "UI / Layout",
    "path_builder":          "UI / Layout",
    "search_layouts":        "UI / Layout",
    "listviews":             "UI / Layout",
    "theme":                 "UI / Layout",
    "component":             "UI / Layout",
    # Reporting / files
    "reports":               "Reporting",
    "dashboard":             "Reporting",
    "file":                  "Files",
    "import_wizard":         "Files",
    # Workflow / automation
    "workflow":              "Workflow",
    "workflow_node":         "Workflow",
    "workflow_edge":         "Workflow",
    "workflow_rules":        "Workflow",
    "process_builders":      "Workflow",
    "flows":                 "Workflow",
    "approval_processes":    "Workflow",
    "node":                  "Workflow",
    "matching_rule":         "Workflow",
    "duplicate_rule":        "Workflow",
    "lead_capture":          "Workflow",
    "email_templates":       "Workflow",
    # Integration / API
    "apex_class":            "Integration",
    "connected_app":         "Integration",
    "named_credential":      "Integration",
    "remote_site_setting":   "Integration",
    "package":               "Integration",
    "webhook":               "Integration",
    "facebook_lead":         "Integration",
    "sf_integration_lead":   "Integration",
    "salesforce_metadata":   "Integration",
    "salesforce_settings":   "Integration",
    "salesforce_sync":       "Integration",
    "landing_numbers":       "Integration / Telephony",
    "telephony_config":      "Integration / Telephony",
    "call_logs":             "Integration / Telephony",
    # Audit / history
    "audit_trails":          "Audit / History",
    "audit_trail_track":     "Audit / History",
    "field_history_log":     "Audit / History",
    "field_tracking_config": "Audit / History",
    "hsitory":               "Audit / History (typo: 'history')",
    "session_log":           "Audit / History",
    "user_login_history":    "Audit / History",
    "group_assignment_tracker":"Audit / History",
    # Messaging
    "message":               "Messaging",
    "channel":               "Messaging",
    # Misc / region / sales / task
    "regions":               "Sales Ops",
    "sales":                 "Sales Ops",
    "task":                  "Sales Ops",
    # Possible legacy duplicates of business-object names
    "customers":             "Legacy / Possible duplicate of 'customer'",
    "products":              "Legacy / Possible duplicate of 'product'",
    "invoices":              "Legacy / Possible duplicate of 'invoice'",
    "invoice_items":         "Legacy / Possible duplicate of 'invoice_item'",
    # Django framework tables
    "django_admin_log":      "Django Framework",
    "django_content_type":   "Django Framework",
    "django_migrations":     "Django Framework",
    "django_session":        "Django Framework",
    "django_celery_beat_clockedschedule":   "Django Framework (Celery Beat)",
    "django_celery_beat_crontabschedule":   "Django Framework (Celery Beat)",
    "django_celery_beat_intervalschedule":  "Django Framework (Celery Beat)",
    "django_celery_beat_periodictask":      "Django Framework (Celery Beat)",
    "django_celery_beat_periodictasks":     "Django Framework (Celery Beat)",
    "django_celery_beat_solarschedule":     "Django Framework (Celery Beat)",
}


# Registry gaps discovered by AST-precise scan of cursor.execute() calls
# and *.sql files. Each entry: (kind, table, label_guess, evidence, recommendation)
REGISTRY_GAPS = [
    # ── Real DB tables used in code that have NO row in the object registry ──
    ("Missing setup table", "homepage_assignment", "Homepage Assignment",
     "api/ORM/setup/newprofile.py:88,93",
     "Add registry row (setup=TRUE); will be migrated to Django ORM"),
    ("Missing setup table", "page_builder", "Page Builder",
     "api/ORM/setup/update_page_builder.py:28",
     "Add registry row (setup=TRUE); distinct from registered "
     "'page_builder_assignment'"),
    ("Missing setup table", "page_component", "Page Component",
     "api/ORM/setup/update_page_builder.py:66,78",
     "Add registry row (setup=TRUE); page-builder child table"),
    ("Missing setup table", "whatsapp_message", "WhatsApp Message",
     "api/BL/whatsapp/utils.py:18,67",
     "Add registry row (setup=TRUE); chat-message store"),
    ("Missing setup table", "user_group_profiles", "User Group Profiles",
     "utils/usergroup_utils.py:305",
     "Add registry row (setup=TRUE); junction table for "
     "user_group <-> profile"),
    ("Missing setup table", "user_group_public_groups",
     "User Group Public Groups",
     "utils/usergroup_utils.py:317; api/workflows/workflow_executor.py:552",
     "Add registry row (setup=TRUE); junction for nested user groups"),
    ("Missing setup table", "otp_verification_sessions",
     "OTP Verification Sessions",
     "public/auth/otp_verification.py:156,198",
     "Add registry row (setup=TRUE); auth/OTP flow"),
    ("Missing setup table", "tenant_provisioning_errors",
     "Tenant Provisioning Errors",
     "public/utils/error_log.py:20",
     "Add registry row (setup=TRUE); tenant onboarding error log"),
    ("Missing setup table", "organizations", "Organizations",
     "adminuser/services/organizations.py:28,53; api/models.py:Organization",
     "Add registry row (setup=TRUE); already has Django model"),

    # ── Tables referenced in mockdata SQL but not registered ──
    ("Missing business object", "report_folder", "Report Folder",
     "sqlfiles/mockdata/reports.sql",
     "Add registry row (setup=FALSE); related to 'reports' (already "
     "registered)"),
    ("Missing business object", "dashboard_component", "Dashboard Component",
     "sqlfiles/mockdata/reports.sql",
     "Add registry row; child of dashboard"),
    ("Missing business object", "dashboard_folders", "Dashboard Folders",
     "sqlfiles/mockdata/reports.sql",
     "Plural form referenced; registry has 'dashboard_folder' singular - "
     "pick one and reconcile"),

    # ── Plural / singular naming inconsistencies ──
    ("Naming inconsistency", "dashboard_folder vs dashboard_folders",
     "Dashboard Folder",
     "registry: dashboard_folder (singular); code: dashboard_folders (plural)",
     "Standardize on one form across registry, code, and mockdata"),
    ("Naming inconsistency", "reports vs report vs report_folder",
     "Report",
     "registry: 'reports'; code uses 'report' and 'report_folder' too",
     "Standardize and register all related tables"),

    # ── Likely typos / bugs ──
    ("Bug — likely typo", "organization (singular)",
     "Probably means 'organizations'",
     "api/pdfgen/views.py:74",
     "Fix the SQL string to use 'organizations'"),
    ("Bug — likely typo", "hsitory",
     "Probably means 'history'",
     "registry row in objects.sql",
     "Rename to 'history' before migration"),

    # ── Legacy duplicates: setup-side vs business-side ──
    ("Legacy duplicate", "customers (setup) vs customer (business)",
     "Customer",
     "objects.sql registers both",
     "Investigate which is actually used; merge or delete the unused one"),
    ("Legacy duplicate", "products (setup) vs product (business)",
     "Product",
     "objects.sql registers both",
     "Investigate which is actually used; merge or delete the unused one"),
    ("Legacy duplicate", "invoices (setup) vs invoice (business)",
     "Invoice",
     "objects.sql registers both",
     "Investigate which is actually used; merge or delete the unused one"),
    ("Legacy duplicate",
     "invoice_items (setup) vs invoice_item (business)",
     "Invoice Item",
     "objects.sql registers both",
     "Investigate which is actually used; merge or delete the unused one"),

    # ── Modeled by Django but no registry entry ──
    ("Django-modeled, not in registry", "facebookleadwebhooks",
     "Facebook Lead Webhooks",
     "facebook/models.py (currently commented out)",
     "Add registry row OR delete the Django model — currently dead code"),

    # ─────────────────────────────────────────────────────────────────
    # The following entries were discovered during the
    # tables.sql / default_tables.sql cross-check (Phase 2 ORM Wave 2).
    # default_tables.sql is loaded by public/utils/organisation.py at
    # tenant provisioning, so every CREATE TABLE in it produces a real
    # per-tenant table — but several of them were never registered in
    # the object table.
    # ─────────────────────────────────────────────────────────────────

    # Real per-tenant tables created by default_tables.sql but absent
    # from the object registry.
    ("Missing setup table", "callactivity", "Call Activity",
     "default_tables.sql:933",
     "Add registry row (setup=TRUE); telephony call-activity log"),
    ("Missing setup table", "dashboard_assignment", "Dashboard Assignment",
     "default_tables.sql (dashboard module)",
     "Add registry row (setup=TRUE); dashboard share assignment"),
    ("Missing setup table", "dashboard_folder_sharing",
     "Dashboard Folder Sharing",
     "default_tables.sql:739",
     "Add registry row (setup=TRUE)"),
    ("Missing setup table", "email_provider_setup",
     "Email Provider Setup",
     "default_tables.sql:785",
     "Add registry row (setup=TRUE); OAuth-token storage (encrypted "
     "since Phase 0.8)"),
    ("Missing setup table", "field_mapping", "Field Mapping",
     "default_tables.sql:146",
     "Add registry row (setup=TRUE)"),
    ("Missing setup table", "notifications", "Notifications",
     "default_tables.sql:905",
     "Add registry row (setup=TRUE); registry currently has BL route "
     "name 'notification' but no entry for the actual table"),
    ("Missing setup table", "org_company", "Organization Company",
     "default_tables.sql (legacy column-mapping table)",
     "Add registry row (setup=TRUE); investigate whether still used"),
    ("Missing setup table", "report_folder_sharing",
     "Report Folder Sharing",
     "default_tables.sql:847",
     "Add registry row (setup=TRUE)"),
    ("Missing setup table", "shared_records",
     "Shared Records (per-record sharing)",
     "default_tables.sql:873",
     "Add registry row (setup=TRUE); DISTINCT from 'sharing_records' "
     "— sharing_records is per-object, shared_records is per-record"),
    ("Missing setup table", "telephony_user", "Telephony User",
     "default_tables.sql:923",
     "Add registry row (setup=TRUE); referenced in api/telephony/views.py"),
    ("Missing setup table", "user_gmail_tokens", "User Gmail Tokens",
     "default_tables.sql:705",
     "Add registry row (setup=TRUE); legacy — superseded by "
     "email_provider_setup but still created"),
    ("Missing setup table", "user_outlook_tokens", "User Outlook Tokens",
     "default_tables.sql:720",
     "Add registry row (setup=TRUE); legacy — superseded by "
     "email_provider_setup but still created"),

    # Naming clashes uncovered by the cross-check.
    ("Naming inconsistency", "report (default_tables.sql) vs reports (registry)",
     "Report",
     "default_tables.sql:320 has `CREATE TABLE report`; registry has "
     "'reports' as the setup name",
     "Pick one: either rename DDL to reports or update the registry "
     "to use 'report'"),
    ("Naming inconsistency",
     "shared_records vs sharing_records",
     "Two genuinely different tables, names too similar",
     "default_tables.sql defines BOTH; sharing_records is per-object, "
     "shared_records is per-record",
     "Rename one to disambiguate (suggest: per_record_shares) and "
     "update all references"),

    # Vestigial registry entries — registered but no DDL anywhere.
    ("Vestigial registry entry", "roles",
     "Aspirational Salesforce-style role hierarchy",
     "Registered as setup=TRUE but no CREATE TABLE in tables.sql / "
     "default_tables.sql / organisation.py / app models; no FROM roles "
     "queries in the codebase",
     "Either build out the role-hierarchy feature OR delete the "
     "registry row + drop the Wave 2 Role model"),
    ("Vestigial registry entry", "owd",
     "Org-Wide Defaults table",
     "Registered as setup=TRUE but no DDL anywhere; no FROM owd queries. "
     "Phase 2 default-deny work uses sharing_records exclusively.",
     "Either build out OR delete the registry row + drop the Wave 2 "
     "OrganizationWideDefault model"),
    ("Vestigial registry entry",
     "apex_class, approval_processes, auth_group, columns_metadata, "
     "component, connected_app, custom_metadata, custom_setting, "
     "duplicate_rule, file, flows, group_assignment_tracker, "
     "import_wizard, lead_capture, lightning_pages, matching_rule, "
     "named_credential, node, package, permission_sets, "
     "process_builders, regions, remote_site_setting, sales, setup, "
     "sf_integration_lead, sharing_rules, tables_metadata, tabs, "
     "theme, users_user_permissions, workflow_rules",
     "Aspirational Salesforce-clone features",
     "Registered as setup=TRUE but no DDL in default_tables.sql, "
     "organisation.py, or app models; no SQL queries reference them",
     "Bulk-investigate; delete vestigial rows after confirming none "
     "are queried at runtime"),

    # Tables.sql discoveries.
    ("Missing business object", "activity",
     "Activity",
     "tables.sql defines `CREATE TABLE activity` but it's not in the "
     "object registry as a business object",
     "Register as business object (setup=FALSE)"),
    ("Missing business object", "email",
     "Email (record)",
     "tables.sql defines `CREATE TABLE email` for sent-email logging; "
     "not in the registry",
     "Register as business object (setup=FALSE)"),
]


# ---------- parsing -------------------------------------------------------- #
COL_ORDER = [
    "id", "allow_activities", "allow_bulk_api_access",
    "allow_in_chatter_groups", "allow_reports", "allow_sharing",
    "allow_streaming_api_access", "datatype", "icon", "icon_color",
    "deployment_status", "description", "enable_licensing", "label",
    "name", "plural_label", "record_name", "search_status",
    "show_tab", "starts_with_vowel_sound", "track_field_history",
    "prefix", "created_date", "last_modified_date",
    "default_access_level", "type", "setup",
]

VALUE_RE = re.compile(r"NULL|'(?:[^'\\]|\\.)*'", re.DOTALL)


# ---------- DDL cross-check sources --------------------------------------- #
TABLES_SQL = os.path.join(ROOT, "tables.sql")
DEFAULT_TABLES_SQL = os.path.join(ROOT, "default_tables.sql")
ORGANISATION_PY = os.path.join(ROOT, "public", "utils", "organisation.py")


_CREATE_TABLE_RE = re.compile(
    r"\bCREATE\s+TABLE\s+(?:IF\s+NOT\s+EXISTS\s+)?"
    r'(?:[a-z_][a-z0-9_]*\.)?\"?([a-z_][a-z0-9_]*)\"?',
    re.IGNORECASE,
)


def extract_create_tables(path: str) -> set[str]:
    if not os.path.exists(path):
        return set()
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return {m.group(1).lower() for m in _CREATE_TABLE_RE.finditer(f.read())}


# Tables that are managed by Django itself or sit in app-specific models.py.
APP_MODELED_DB_TABLES: set[str] = {
    # api/models.py
    "organizations", "users", "session_log", "user_login_history",
    # facebook/
    "facebook_lead", "facebookleadwebhooks",
    # whatsapp/
    "message", "channel", "webhook",
    # sf_integration/
    "salesforce_metadata", "salesforce_settings", "salesforce_sync",
}

# Tables modeled in Phase 2 Wave 2 (api/tenant_models/).
WAVE2_DB_TABLES: set[str] = {
    "object", "fields", "profile", "roles",
    "user_group", "user_group_users",
    "object_permissions", "field_permissions",
    "tab_permissions", "app_permissions",
    "sharing_records", "owd",
}

# Tables created by Python f-string DDL inside organisation.py (the
# tenant-provisioning bootstrap pair).
ORG_PY_INLINE_DDL: set[str] = {"profile", "users"}


def parse_objects_sql(path: str):
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()

    start = text.find('INSERT INTO "object"')
    if start < 0:
        sys.exit("Could not find INSERT INTO \"object\" in objects.sql")
    block = text[start:]

    # naive paren-balanced row extractor
    rows = []
    depth = 0
    buf = []
    in_string = False
    for ch in block:
        if ch == "'" and (not buf or buf[-1] != "\\"):
            in_string = not in_string
        if not in_string:
            if ch == "(":
                if depth == 0:
                    buf = []
                depth += 1
            elif ch == ")":
                depth -= 1
                if depth == 0:
                    rows.append("(" + "".join(buf) + ")")
                    continue
        if depth > 0:
            buf.append(ch)

    parsed = []
    for raw in rows:
        values = VALUE_RE.findall(raw)
        if len(values) < len(COL_ORDER):
            continue
        rec = {}
        for col, val in zip(COL_ORDER, values):
            if val == "NULL":
                rec[col] = None
            else:
                rec[col] = val.strip("'")
        parsed.append(rec)
    return parsed


# ---------- categorization ------------------------------------------------- #
def categorize(rec):
    setup = rec.get("setup") == "True"
    typ = (rec.get("type") or "").lower()
    if setup:
        return "SETUP"
    if typ == "custom":
        return "BUSINESS / CUSTOM"
    return "BUSINESS / STANDARD"


def subcategory(rec):
    name = rec["name"]
    if rec.get("setup") == "True":
        return SETUP_GROUPS.get(name, "Setup (uncategorized)")
    if (rec.get("type") or "").lower() == "custom":
        return "User-defined"
    return "Standard business object"


def migration_recommendation(rec):
    name = rec["name"]
    if rec.get("setup") == "True":
        if name.startswith("django_"):
            return "Skip — managed by Django/Celery framework"
        if name in DJANGO_MODELED:
            return "Already in Django ORM — audit & harden"
        if name == "hsitory":
            return "Rename to 'history' then migrate to Django ORM"
        if name in ("customers", "products", "invoices", "invoice_items"):
            return "Investigate — likely legacy duplicate of business object"
        return "Migrate to Django ORM"
    return "Keep raw SQL — dynamic schema via metadata gateway"


# ---------- workbook building --------------------------------------------- #
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SETUP_FILL = PatternFill("solid", fgColor="DDEBF7")
STD_BIZ_FILL = PatternFill("solid", fgColor="FFF2CC")
CUSTOM_BIZ_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"


def autosize(ws, headers, max_w=60):
    for i, h in enumerate(headers, start=1):
        col_letter = get_column_letter(i)
        max_len = len(str(h))
        for cell in ws[col_letter][1:]:
            v = cell.value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_w, max_len + 2)


def add_summary(wb, parsed):
    ws = wb.create_sheet("Summary", 0)
    ws.append(["Bucket", "Count", "Notes"])

    n_setup = sum(1 for r in parsed if r.get("setup") == "True")
    n_std = sum(1 for r in parsed if r.get("setup") != "True"
                and (r.get("type") or "").lower() == "standard")
    n_cust = sum(1 for r in parsed if r.get("setup") != "True"
                 and (r.get("type") or "").lower() == "custom")
    n_django_modeled = sum(1 for r in parsed if r["name"] in DJANGO_MODELED)
    n_django_framework = sum(1 for r in parsed if r["name"].startswith("django_"))
    n_legacy = sum(1 for r in parsed if r["name"]
                   in ("customers", "products", "invoices", "invoice_items"))

    rows = [
        ("Total objects in registry", len(parsed), ""),
        ("Setup objects (setup=TRUE)", n_setup, "Fixed schema → migrate to Django ORM"),
        ("Standard business objects (setup=FALSE, type=standard)", n_std,
         "Dynamic schema → keep raw SQL via gateway"),
        ("Custom business objects (setup=FALSE, type=custom)", n_cust,
         "User-defined → keep raw SQL via gateway"),
        ("", "", ""),
        ("Already in Django ORM (db_table mapped)", n_django_modeled,
         "12 setup tables — audit & harden"),
        ("Setup tables that are Django framework / Celery", n_django_framework,
         "Skip — managed by Django/Celery"),
        ("Setup tables suspected legacy duplicates of business objects",
         n_legacy, "customers/products/invoices/invoice_items"),
        ("", "", ""),
        ("Setup tables remaining to migrate to Django ORM",
         n_setup - n_django_modeled - n_django_framework - n_legacy,
         "Primary migration scope"),
    ]
    for r in rows:
        ws.append(r)
    style_header(ws, 3)
    autosize(ws, ["Bucket", "Count", "Notes"])


def add_all_objects(wb, parsed):
    ws = wb.create_sheet("All_Objects")
    headers = [
        "Category", "Subcategory", "Name", "Label", "Plural Label",
        "Type", "Setup?", "Default Access Level", "Deployment Status",
        "Show Tab", "Allow Reports", "Allow Sharing",
        "Track Field History", "Prefix", "Currently in Django ORM?",
        "Django Model Source", "Migration Recommendation",
    ]
    ws.append(headers)

    # sort: SETUP first, then STANDARD, then CUSTOM; within each by name
    cat_order = {"SETUP": 0, "BUSINESS / STANDARD": 1, "BUSINESS / CUSTOM": 2}
    sorted_recs = sorted(parsed, key=lambda r: (cat_order[categorize(r)], r["name"]))

    for r in sorted_recs:
        cat = categorize(r)
        in_django = r["name"] in DJANGO_MODELED
        ws.append([
            cat,
            subcategory(r),
            r["name"],
            r.get("label"),
            r.get("plural_label"),
            r.get("type"),
            r.get("setup"),
            r.get("default_access_level"),
            r.get("deployment_status"),
            r.get("show_tab"),
            r.get("allow_reports"),
            r.get("allow_sharing"),
            r.get("track_field_history"),
            r.get("prefix"),
            "Yes" if in_django else "No",
            DJANGO_MODELED.get(r["name"], ""),
            migration_recommendation(r),
        ])

    style_header(ws, len(headers))

    # color-band rows by category
    for i, r in enumerate(sorted_recs, start=2):
        cat = categorize(r)
        fill = (SETUP_FILL if cat == "SETUP"
                else STD_BIZ_FILL if cat == "BUSINESS / STANDARD"
                else CUSTOM_BIZ_FILL)
        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).fill = fill
            ws.cell(row=i, column=c).border = BORDER
            ws.cell(row=i, column=c).alignment = Alignment(
                vertical="top", wrap_text=True
            )

    autosize(ws, headers)

    # turn it into a real Excel table for filtering
    last_col = get_column_letter(len(headers))
    last_row = ws.max_row
    table = Table(displayName="AllObjects", ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showRowStripes=False, showColumnStripes=False
    )
    ws.add_table(table)


def add_migration_plan(wb, parsed):
    ws = wb.create_sheet("Migration_Plan")
    headers = [
        "Phase", "Wave", "Object Name", "Label", "Subcategory",
        "Action", "Notes",
    ]
    ws.append(headers)

    waves = OrderedDict()
    waves["Wave 1 — Core platform metadata"] = [
        "object", "fields", "tables_metadata", "columns_metadata", "setup",
    ]
    waves["Wave 2 — Authorization (high-risk)"] = [
        "profile", "permission_sets", "roles",
        "user_group", "user_group_users", "users_user_permissions",
        "object_permissions", "field_permissions",
        "tab_permissions", "app_permissions",
        "sharing_records", "sharing_rules", "owd",
    ]
    waves["Wave 3 — UI / layout / navigation"] = [
        "app", "tabs", "page_layouts", "page_builder_assignment",
        "layout_assignment", "lightning_pages", "path_builder",
        "search_layouts", "listviews", "theme", "component",
    ]
    waves["Wave 4 — Reporting / files"] = [
        "reports", "dashboard", "file", "import_wizard",
    ]
    waves["Wave 5 — Workflow / automation"] = [
        "workflow", "workflow_node", "workflow_edge", "workflow_rules",
        "process_builders", "flows", "approval_processes", "node",
        "matching_rule", "duplicate_rule", "lead_capture", "email_templates",
    ]
    waves["Wave 6 — Integration / telephony"] = [
        "apex_class", "connected_app", "named_credential", "remote_site_setting",
        "package", "sf_integration_lead",
        "landing_numbers", "telephony_config", "call_logs",
    ]
    waves["Wave 7 — Audit / history"] = [
        "audit_trails", "audit_trail_track", "field_history_log",
        "field_tracking_config", "hsitory", "group_assignment_tracker",
    ]
    waves["Wave 8 — Misc / sales ops"] = [
        "regions", "sales", "task", "custom_metadata", "custom_setting",
    ]

    name_to_rec = {r["name"]: r for r in parsed}

    # Phase 1 — already-modeled audit
    for n in sorted(DJANGO_MODELED.keys()):
        rec = name_to_rec.get(n)
        if not rec:
            continue
        ws.append([
            "Phase 1 — Audit existing Django models",
            "—",
            n,
            rec.get("label"),
            subcategory(rec),
            "Audit & harden",
            DJANGO_MODELED[n],
        ])

    # Phase 2 — wave-based migration of remaining setup tables
    for wave_name, objs in waves.items():
        for n in objs:
            rec = name_to_rec.get(n)
            if not rec:
                continue
            if n in DJANGO_MODELED:
                continue
            ws.append([
                "Phase 2 — Migrate setup tables to Django ORM",
                wave_name,
                n,
                rec.get("label"),
                subcategory(rec),
                "Build Django model + migration; replace raw SQL callsites",
                "",
            ])

    # Phase 3 — investigate legacy duplicates
    for n in ("customers", "products", "invoices", "invoice_items"):
        rec = name_to_rec.get(n)
        if not rec:
            continue
        ws.append([
            "Phase 3 — Resolve legacy duplicates",
            "—",
            n,
            rec.get("label"),
            subcategory(rec),
            "Investigate & merge or delete",
            f"Possible duplicate of business object '{n.rstrip('s')}'",
        ])

    # Phase 4 — Django framework — leave alone
    for n in sorted(r["name"] for r in parsed if r["name"].startswith("django_")):
        rec = name_to_rec.get(n)
        if not rec:
            continue
        ws.append([
            "Phase 4 — Skip (Django/Celery managed)",
            "—",
            n,
            rec.get("label"),
            subcategory(rec),
            "No action — framework owns it",
            "",
        ])

    # Phase 5 — Business objects — keep raw SQL
    for r in sorted(parsed, key=lambda x: x["name"]):
        if r.get("setup") == "True":
            continue
        ws.append([
            "Phase 5 — Business objects (raw SQL via gateway)",
            "Standard" if (r.get("type") or "").lower() == "standard" else "Custom",
            r["name"],
            r.get("label"),
            subcategory(r),
            "Route through hardened ORM gateway (sql.Identifier + schema authority)",
            "Dynamic schema — cannot be a static Django model",
        ])

    style_header(ws, len(headers))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                            min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER

    autosize(ws, headers, max_w=80)

    last_col = get_column_letter(len(headers))
    table = Table(displayName="MigrationPlan",
                  ref=f"A1:{last_col}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9", showRowStripes=True
    )
    ws.add_table(table)


# Color fills per gap kind for the Registry_Gaps sheet
GAP_FILLS = {
    "Missing setup table":             PatternFill("solid", fgColor="F8CBAD"),
    "Missing business object":         PatternFill("solid", fgColor="FCE4D6"),
    "Naming inconsistency":            PatternFill("solid", fgColor="FFF2CC"),
    "Bug — likely typo":               PatternFill("solid", fgColor="F4B084"),
    "Legacy duplicate":                PatternFill("solid", fgColor="DDEBF7"),
    "Django-modeled, not in registry": PatternFill("solid", fgColor="E2EFDA"),
}


def add_registry_gaps(wb):
    ws = wb.create_sheet("Registry_Gaps")
    headers = ["Kind", "Table / Item", "Label / Note",
               "Evidence (file:line)", "Recommended Action"]
    ws.append(headers)

    # group by kind, preserve insertion order
    order = ["Missing setup table", "Missing business object",
             "Naming inconsistency", "Bug — likely typo",
             "Legacy duplicate", "Django-modeled, not in registry"]
    rows = sorted(REGISTRY_GAPS,
                  key=lambda r: (order.index(r[0]) if r[0] in order else 99,
                                 r[1]))
    for r in rows:
        ws.append(list(r))

    style_header(ws, len(headers))
    for i, r in enumerate(rows, start=2):
        fill = GAP_FILLS.get(r[0])
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=i, column=c)
            if fill is not None:
                cell.fill = fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    autosize(ws, headers, max_w=80)

    last_col = get_column_letter(len(headers))
    table = Table(displayName="RegistryGaps",
                  ref=f"A1:{last_col}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight16", showRowStripes=False
    )
    ws.add_table(table)


def update_summary_with_gaps(wb):
    """Append registry-gap counts to the Summary sheet."""
    ws = wb["Summary"]
    ws.append(("", "", ""))
    ws.append(("REGISTRY GAPS (from AST scan + .sql review)", "", ""))
    by_kind = {}
    for kind, *_ in REGISTRY_GAPS:
        by_kind[kind] = by_kind.get(kind, 0) + 1
    for kind in ("Missing setup table", "Missing business object",
                 "Naming inconsistency", "Bug — likely typo",
                 "Legacy duplicate", "Django-modeled, not in registry"):
        if kind in by_kind:
            ws.append((kind, by_kind[kind], ""))
    ws.append(("Total registry gaps", len(REGISTRY_GAPS),
               "See Registry_Gaps sheet"))


def add_ddl_cross_check(wb, parsed):
    """Compare every distinct table name across all sources of truth.

    Sources:
      A. ``sqlfiles/objects.sql`` registry (parsed)
      B. ``default_tables.sql`` (per-tenant setup-table DDL)
      C. ``tables.sql`` (per-tenant business-object DDL)
      D. ``public/utils/organisation.py`` (inline DDL — bootstrap tables)
      E. App-modeled tables (Django models in app/models.py)
      F. Phase 2 Wave 2 Django models

    The sheet shows, for every distinct table name, whether each source
    has it. Reviewers can scan for mismatches in seconds.
    """
    ws = wb.create_sheet("DDL_Cross_Check")

    headers = [
        "Table Name",
        "In Registry",
        "Registry Type",        # setup / standard / custom / -
        "default_tables.sql",   # canonical per-tenant setup DDL
        "tables.sql",           # canonical per-tenant business DDL
        "Inline organisation.py",
        "App models.py",
        "Phase 2 Wave 2 model",
        "Verdict",
    ]
    ws.append(headers)

    default_set = extract_create_tables(DEFAULT_TABLES_SQL)
    tables_set = extract_create_tables(TABLES_SQL)

    registry_setup = {r["name"] for r in parsed if r.get("setup") == "True"}
    registry_business_std = {r["name"] for r in parsed
                             if r.get("setup") != "True"
                             and (r.get("type") or "").lower() == "standard"}
    registry_business_cust = {r["name"] for r in parsed
                              if r.get("setup") != "True"
                              and (r.get("type") or "").lower() == "custom"}

    def _registry_type(name):
        if name in registry_setup:
            return "setup"
        if name in registry_business_std:
            return "standard"
        if name in registry_business_cust:
            return "custom"
        return "-"

    # Union of all table names from every source.
    all_names = sorted(
        default_set
        | tables_set
        | ORG_PY_INLINE_DDL
        | APP_MODELED_DB_TABLES
        | WAVE2_DB_TABLES
        | registry_setup
        | registry_business_std
        | registry_business_cust
    )

    for name in all_names:
        in_registry = name in (
            registry_setup | registry_business_std | registry_business_cust
        )
        in_default = name in default_set
        in_tables = name in tables_set
        in_org_py = name in ORG_PY_INLINE_DDL
        in_app_models = name in APP_MODELED_DB_TABLES
        in_wave2 = name in WAVE2_DB_TABLES

        ddl_sources = sum([in_default, in_tables, in_org_py, in_app_models])

        # Compute verdict.
        if name.startswith("django_"):
            verdict = "OK — Django/Celery framework table"
        elif in_registry and ddl_sources >= 1:
            verdict = "OK"
        elif in_registry and ddl_sources == 0:
            verdict = "VESTIGIAL — registered but no DDL anywhere"
        elif (not in_registry) and ddl_sources >= 1:
            verdict = "GAP — DDL exists, registry row missing"
        else:
            verdict = "?"

        ws.append([
            name,
            "Yes" if in_registry else "No",
            _registry_type(name),
            "Yes" if in_default else "",
            "Yes" if in_tables else "",
            "Yes" if in_org_py else "",
            "Yes" if in_app_models else "",
            "Yes" if in_wave2 else "",
            verdict,
        ])

    style_header(ws, len(headers))

    # Color-code rows by verdict.
    verdict_fills = {
        "OK": PatternFill("solid", fgColor="E2EFDA"),
        "OK — Django/Celery framework table": PatternFill("solid", fgColor="EAEAEA"),
        "GAP — DDL exists, registry row missing": PatternFill("solid", fgColor="FCE4D6"),
        "VESTIGIAL — registered but no DDL anywhere": PatternFill("solid", fgColor="FFF2CC"),
    }
    for i, name in enumerate(all_names, start=2):
        verdict = ws.cell(row=i, column=len(headers)).value
        fill = verdict_fills.get(verdict)
        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(row=i, column=c).fill = fill
        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).border = BORDER
            ws.cell(row=i, column=c).alignment = Alignment(
                vertical="top", wrap_text=True
            )

    autosize(ws, headers, max_w=60)

    last_col = get_column_letter(len(headers))
    table = Table(displayName="DDLCrossCheck",
                  ref=f"A1:{last_col}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight2", showRowStripes=False, showColumnStripes=False
    )
    ws.add_table(table)

    # Summary block at the bottom.
    counts = {
        "OK": 0,
        "OK — Django/Celery framework table": 0,
        "GAP — DDL exists, registry row missing": 0,
        "VESTIGIAL — registered but no DDL anywhere": 0,
    }
    for name in all_names:
        if name.startswith("django_"):
            counts["OK — Django/Celery framework table"] += 1
            continue
        in_registry = name in (
            registry_setup | registry_business_std | registry_business_cust
        )
        ddl_sources = sum([
            name in default_set, name in tables_set,
            name in ORG_PY_INLINE_DDL, name in APP_MODELED_DB_TABLES,
        ])
        if in_registry and ddl_sources >= 1:
            counts["OK"] += 1
        elif in_registry and ddl_sources == 0:
            counts["VESTIGIAL — registered but no DDL anywhere"] += 1
        elif not in_registry and ddl_sources >= 1:
            counts["GAP — DDL exists, registry row missing"] += 1


def main():
    parsed = parse_objects_sql(OBJECTS_SQL)
    print(f"Parsed {len(parsed)} object rows")

    wb = Workbook()
    # remove default sheet
    default = wb.active
    wb.remove(default)

    add_summary(wb, parsed)
    add_all_objects(wb, parsed)
    add_migration_plan(wb, parsed)
    add_registry_gaps(wb)
    add_ddl_cross_check(wb, parsed)
    update_summary_with_gaps(wb)

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}")


if __name__ == "__main__":
    main()
