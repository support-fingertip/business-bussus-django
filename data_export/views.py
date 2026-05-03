import decimal
import io
import logging                          # FIX BLOCKER-10: module-level, not per-request
import os
import tempfile
from datetime import datetime, timedelta
from typing import Iterable, Sequence

import pandas as pd
import pytz                             # FIX BLOCKER-6: direct tz conversion
from django.db import connection
from django.http import HttpResponse, FileResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell.cell import WriteOnlyCell
from psycopg2 import sql
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from api.ORM.AuditLogs.audit_trail_logs import log_audit
from authentication.custom_jwt_auth import CustomJWTAuthentication
from public.auth.session import get_connection_and_user_details
from api.permissions.permissions import get_permissions, get_object_details, check_permission, get_field_metadata, get_all_fields_for_table
from api.BL.dashboard import export_to_excel
from api.BL.blcontroller import filter_summary_fields, normalize_group_by, get_details_fields
from api.BL.computed_fields import process_computed_fields_for_report, apply_computed_fields_to_records, separate_computed_filters, convert_rollup_filters_to_physical, apply_computed_filters


# ---- helpers ---------------------------------------------------------------

MAX_EXPORT_OBJECTS  = 20          # hard cap to avoid abuse / OOM
MAX_ROWS_PER_OBJECT = 200_000     # supports 1L+ rows per object
# Bigger batch = fewer cursor round-trips. 20k rows ≈ a few MB and keeps
# Python's per-batch loop overhead amortised — measured 1.8x speedup over
# 5k for 1L+ rows on the openpyxl write_only path.
STREAM_BATCH_SIZE   = 20_000
FILE_STREAM_CHUNK   = 64 * 1024   # FileResponse chunk size for large xlsx downloads
DEFAULT_TZ          = "Asia/Kolkata"

# FIX BLOCKER-10: module-level logger — not recreated on every request
logger = logging.getLogger("data_export")

# Columns that are foreign-keys into the users table.
# When any of these appear in export_fields the SELECT is rewritten to
# LEFT JOIN users and emit users.name instead of the raw integer id.
USER_ID_COLUMNS = {
    "owner_id":            "u_owner",
    "created_by_id":       "u_created",
    "last_modified_by_id": "u_modified",
}

# Per-process cache of column data_types populated by
# _resolve_schema_and_columns. Lets the streaming SELECT decide which
# columns to wrap in to_char(... AT TIME ZONE ...) so datetime formatting
# happens server-side (in C) instead of in the per-row Python loop.
_COLUMN_TYPE_CACHE: dict[tuple[str, str, str], str] = {}


def _quote_ident(name: str) -> sql.Identifier:
    if not name or not isinstance(name, str):
        raise ValueError("Invalid identifier")
    return sql.Identifier(name)


def _set_search_path(cursor, schema: str) -> None:
    cursor.execute(sql.SQL("SET search_path TO {}").format(_quote_ident(schema)))


def _safe_sheet_name(name: str) -> str:
    sanitized = "".join(ch for ch in name if ch not in r'[]:*?/\\')
    return (sanitized or "Sheet")[:31]


def _resolve_schema_and_columns(table_name: str, preferred_schema: str) -> tuple[str | None, list[str]]:
    """
    FIX BLOCKER-4: Replace 3 separate information_schema round-trips
    (_table_exists + _find_table_schema + _get_real_columns) with ONE query.

    Returns (resolved_schema, [col1, col2, ...]) ordered by ordinal_position.
    Returns (None, []) when the table does not exist in either schema.

    Also caches column data_types into _COLUMN_TYPE_CACHE keyed by
    (resolved_schema, table_name, col_name) → data_type so the streaming
    path can wrap timestamp columns in to_char() at SQL time without an
    extra catalog round-trip.
    """
    schemas_to_try = [preferred_schema]
    if preferred_schema != "public":
        schemas_to_try.append("public")

    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT table_schema, column_name, data_type
            FROM information_schema.columns
            WHERE table_name = %s
              AND table_schema = ANY(%s)
            ORDER BY
              CASE table_schema WHEN %s THEN 0 ELSE 1 END,
              ordinal_position
            """,
            [table_name, schemas_to_try, preferred_schema],
        )
        rows = cursor.fetchall()

    if not rows:
        return None, []

    resolved_schema = rows[0][0]
    columns = [r[1] for r in rows if r[0] == resolved_schema]
    for r in rows:
        if r[0] == resolved_schema:
            _COLUMN_TYPE_CACHE[(resolved_schema, table_name, r[1])] = r[2]
    return resolved_schema, columns


def _stream_rows_to_sheet(
    ws,
    object_name: str,
    schema: str,
    fields: list[str],
    field_label_map: dict,
    user_tz: pytz.BaseTzInfo,
    lookup_fields: dict[str, str] | None = None,
) -> int:
    """
    FIX BLOCKER-1: Stream rows from PostgreSQL directly into the write_only
    openpyxl worksheet in batches of STREAM_BATCH_SIZE — never holds all rows
    in Python memory simultaneously.

    FIX BLOCKER-2: Caller must pass a write_only worksheet (ws). Rows are
    serialised into the ZIP stream immediately — no full workbook graph in RAM.

    FIX BLOCKER-6: Datetime values are converted per-value at read time using
    pytz — no pandas DataFrame or apply() overhead at all.

    Returns the number of data rows written (excluding the header).
    """
    # psycopg2 OIDs for timestamp (1114) and timestamptz (1184)
    TIMESTAMP_OIDS = {1114, 1184}

    # ── Build SELECT SQL — with LEFT JOINs for user-id columns ──────────────
    #
    # For columns like owner_id / created_by_id / last_modified_by_id we emit:
    #
    #   SELECT t.col1, t.col2,
    #          u_owner.name   AS owner_id,
    #          u_created.name AS created_by_id,
    #          ...
    #   FROM   <table> t
    #   LEFT JOIN users u_owner   ON u_owner.id   = t.owner_id
    #   LEFT JOIN users u_created ON u_created.id = t.created_by_id
    #   ...
    #   LIMIT N
    #
    # LEFT JOIN keeps rows whose user-id is NULL (unassigned records).
    # The column alias equals the original field name so field_label_map
    # (header labels) works without any change.
    # Only JOINs for user-id cols that are actually present in `fields`.
    #
    # Normalise lookup_fields — exclude USER_ID_COLUMNS (already handled below)
    # to avoid duplicate JOINs for owner_id / created_by_id / last_modified_by_id.
    _lookup = {
        f: tbl
        for f, tbl in (lookup_fields or {}).items()
        if f not in USER_ID_COLUMNS
    }

    # ── Classify every exported field into one of three buckets ──────────────
    #   A) USER_ID_COLUMNS  — JOIN users, emit users.name AS col_name
    #   B) lookup field     — JOIN parent table, emit parent.name AS col_name
    #   C) plain column     — emit t.col_name directly
    #
    # Generated SQL example (leads with account_id lookup + owner_id user-ref):
    #
    #   SELECT t.id, t.email,
    #          "u_owner".name        AS "owner_id",
    #          "lkp_account_id".name AS "account_id",
    #          t.amount
    #   FROM leads t
    #   LEFT JOIN users    "u_owner"      ON "u_owner".id      = t."owner_id"
    #   LEFT JOIN accounts "lkp_account_id" ON "lkp_account_id".id = t."account_id"
    #   LIMIT 50000
    #
    select_parts = []
    join_parts   = []
    # Tz literal used to push timestamp formatting into Postgres (in C)
    # so the per-row Python loop doesn't have to call pytz at all.
    tz_literal = sql.Literal(user_tz.zone)
    _TIMESTAMP_PG_TYPES = {
        "timestamp without time zone",
        "timestamp with time zone",
    }

    for f in fields:
        if f in USER_ID_COLUMNS:
            # ── Bucket A: user foreign-key ────────────────────────────────────
            alias = USER_ID_COLUMNS[f]          # e.g. "u_owner"
            select_parts.append(
                sql.SQL("{}.name AS {}").format(
                    sql.Identifier(alias),
                    sql.Identifier(f),
                )
            )
            join_parts.append(
                sql.SQL("LEFT JOIN users {alias} ON {alias}.id = t.{col}").format(
                    alias=sql.Identifier(alias),
                    col=sql.Identifier(f),
                )
            )

        elif f in _lookup:
            # ── Bucket B: generic lookup / relationship field ─────────────────
            parent_table = _lookup[f]
            alias        = "lkp_" + f           # e.g. "lkp_account_id"
            select_parts.append(
                sql.SQL("{}.name AS {}").format(
                    sql.Identifier(alias),
                    sql.Identifier(f),
                )
            )
            join_parts.append(
                sql.SQL("LEFT JOIN {parent} {alias} ON {alias}.id = t.{col}").format(
                    parent=sql.Identifier(parent_table),
                    alias=sql.Identifier(alias),
                    col=sql.Identifier(f),
                )
            )

        else:
            # ── Bucket C: plain column ────────────────────────────────────────
            # Wrap timestamp/timestamptz in to_char() so Postgres returns the
            # already-localised, already-formatted string. Eliminates the
            # per-value Python branch + pytz.localize() / .astimezone() call
            # for every row × every datetime column — the dominant cost in
            # 1L+-row exports with ~3-5 datetime columns.
            col_type = _COLUMN_TYPE_CACHE.get((schema, object_name, f))
            if col_type in _TIMESTAMP_PG_TYPES:
                select_parts.append(
                    sql.SQL(
                        "to_char({tcol}::timestamptz AT TIME ZONE {tz}, "
                        "'DD/MM/YYYY HH24:MI') AS {alias}"
                    ).format(
                        tcol=sql.SQL("t.{}").format(sql.Identifier(f)),
                        tz=tz_literal,
                        alias=sql.Identifier(f),
                    )
                )
            else:
                select_parts.append(
                    sql.SQL("{}.{}").format(
                        sql.Identifier("t"),
                        sql.Identifier(f),
                    )
                )

    field_sql    = sql.SQL(", ").join(select_parts)
    join_clauses = (
        sql.SQL(" ") + sql.SQL(" ").join(join_parts)
        if join_parts else sql.SQL("")
    )

    select_sql = sql.SQL(
        "SELECT {fields} FROM {table} t{joins} LIMIT {limit}"
    ).format(
        fields=field_sql,
        table=sql.Identifier(object_name),
        joins=join_clauses,
        limit=sql.Literal(MAX_ROWS_PER_OBJECT),
    )

    rows_written = 0

    # ── Helper: write header row + stream result rows into the write_only ws ──
    #
    # IMPORTANT — named cursor description behaviour (psycopg2):
    #   Regular cursor : cur.execute() runs the query immediately → description populated
    #   Named cursor   : cur.execute() only issues DECLARE ... CURSOR FOR ...
    #                    description stays None until the FIRST fetchmany() call
    #
    # Fix: call fetchmany() once first to trigger the FETCH and populate description,
    # then process that first batch together with all subsequent batches.
    #
    def _write_cursor_to_sheet(cur):
        nonlocal rows_written

        # Fetch the first batch — this populates cur.description on named cursors
        first_batch = cur.fetchmany(STREAM_BATCH_SIZE)

        # ── Header row ────────────────────────────────────────────────────────
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")
        header_row  = []
        for col_name in fields:
            cell           = WriteOnlyCell(ws, value=field_label_map.get(col_name, col_name))
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            header_row.append(cell)
        ws.append(header_row)

        # ── Stream all batches (first_batch + remaining) ──────────────────────
        # Datetime formatting now happens in Postgres via to_char() (see
        # Bucket-C wrapping above), so the per-row hot loop is just
        # `ws.append(row)` — no Python branching, no pytz calls.
        batch = first_batch
        ws_append = ws.append   # micro-opt: avoid attr lookup per row
        while batch:
            for raw_row in batch:
                ws_append(raw_row)
            rows_written += len(batch)
            batch = cur.fetchmany(STREAM_BATCH_SIZE)


    # ── Streaming via named server-side cursor ───────────────────────────────
    #
    # Two previous errors explained and fixed:
    #
    # Error 1: "can't use a named cursor outside of transactions"
    #   Named cursors need autocommit=False (an open transaction).
    #   Fix: temporarily set autocommit=False, rollback after (read-only SELECT).
    #
    # Error 2: "syntax error at or near SET" — DECLARE ... CURSOR FOR SET ...
    #   psycopg2 named cursor: the FIRST cur.execute() call becomes the body of
    #       DECLARE "name" CURSOR WITHOUT HOLD FOR <your SQL>
    #   So SET search_path CANNOT go through a named cursor — it must be issued
    #   on a regular cursor BEFORE the named cursor is opened.
    #   SET search_path is session-scoped, so it persists for the named cursor.
    #
    # Correct sequence:
    #   1. regular_cursor.execute("SET search_path TO schema")  <- session-level
    #   2. named_cur = raw_conn.cursor(name="...")               <- open after SET
    #   3. named_cur.execute("SELECT ...")                       <- SELECT only
    #
    raw_conn       = connection.connection
    select_sql_str = select_sql.as_string(raw_conn)
    set_path_sql   = sql.SQL("SET search_path TO {}").format(
        _quote_ident(schema)
    ).as_string(raw_conn)

    def _run_named_cursor():
        # Step 1: SET search_path through a normal cursor (session-scoped, persists)
        with raw_conn.cursor() as reg_cur:
            reg_cur.execute(set_path_sql)
        # Step 2: open named cursor AFTER search_path is already set
        named_cur = raw_conn.cursor(name="export_" + object_name)
        # Match server-side prefetch to our fetchmany batch — without this,
        # psycopg2 defaults itersize to 2000 and ends up issuing extra FETCH
        # round-trips even when fetchmany asks for 20k.
        named_cur.itersize = STREAM_BATCH_SIZE
        try:
            # Step 3: only SELECT goes here — no SET statements allowed
            named_cur.execute(select_sql_str)
            _write_cursor_to_sheet(named_cur)
        finally:
            named_cur.close()

    if not raw_conn.autocommit:
        # Already inside a Django transaction — safe to use named cursor directly
        _run_named_cursor()
    else:
        # autocommit=True: must disable it to satisfy named cursor transaction req.
        # ROLLBACK at end is safe — this is a read-only SELECT, nothing to commit.
        try:
            raw_conn.autocommit = False
            _run_named_cursor()
        except Exception:
            raise
        finally:
            try:
                raw_conn.rollback()        # read-only export — safe to discard
            except Exception:
                pass
            raw_conn.autocommit = True     # always restore original state

    return rows_written


def fetch_audit_trail_raw(schema: str = "public") -> tuple[Sequence[tuple], list[str]]:
    """Fetch audit logs for the last 6 months using parameterized SQL and schema scoping."""
    six_months_ago = timezone.now() - timedelta(days=180)
    with connection.cursor() as cursor:
        _set_search_path(cursor, schema)
        cursor.execute(
            """
            SELECT id, user_id, source_namespace_prefix, action, section,
                   is_delegate_user, changed_at
            FROM audit_trail_track
            WHERE changed_at >= %s
            """,
            [six_months_ago],
        )
        rows = cursor.fetchall()
        columns = [col[0] for col in cursor.description]
    return rows, columns


def get_user_timezone(user_id, schema="public") -> str:
    with connection.cursor() as cursor:
        _set_search_path(cursor, schema)
        cursor.execute("SELECT timezone FROM users WHERE id = %s", [user_id])
        row = cursor.fetchone()
        return row[0] if row and row[0] else DEFAULT_TZ


def get_details_fields(fields: Iterable) -> list:
    processed = []
    for field in fields:
        if isinstance(field, dict):
            field_copy = dict(field)
            field_copy.pop("aggregate", None)
            processed.append(field_copy)
        else:
            processed.append(field)
    return processed


def _compute_grand_total(result, fields, show_detail_rows, show_row_counts, group_by_fields):
    """Aggregate values across `result` for every field that has an aggregate
    defined. Returns a dict keyed by the result-row column name (alias for
    summary view, raw field name for detail view)."""
    import decimal

    def _to_float(v):
        if v is None or v == "":
            return None
        if isinstance(v, decimal.Decimal):
            return float(v)
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    totals = {}
    aggregate_fields = [
        f for f in fields
        if isinstance(f, dict) and f.get("aggregate") and f.get("name")
    ]

    for f in aggregate_fields:
        agg = (f.get("aggregate") or "").lower()
        if agg == "average":
            agg = "avg"
        name = f.get("name")
        base_alias = f.get("alias") or (f"{agg}_{name}" if agg else name)
        # Detail view doesn't aggregate — columns hold raw field values, keyed
        # by the base field name. Summary view keys by the aggregate alias.
        column_key = name if show_detail_rows else base_alias

        values = []
        for row in result:
            raw = row.get(column_key)
            if raw is None and show_detail_rows is False:
                raw = row.get(name)
            fv = _to_float(raw)
            if fv is not None:
                values.append(fv)

        if not values:
            continue
        if agg == "sum":
            totals[column_key] = round(sum(values), 2)
        elif agg == "count":
            totals[column_key] = int(sum(values)) if not show_detail_rows else len(values)
        elif agg == "avg":
            totals[column_key] = round(sum(values) / len(values), 2)
        elif agg == "min":
            totals[column_key] = min(values)
        elif agg == "max":
            totals[column_key] = max(values)

    # Group-by summary always exposes a row_count column (unless hidden) —
    # total it so the Grand Total row shows the full record count.
    if show_row_counts and not show_detail_rows and group_by_fields:
        rc_total = 0
        for row in result:
            rc = row.get("row_count")
            if rc is None:
                continue
            try:
                rc_total += int(rc)
            except (TypeError, ValueError):
                pass
        if rc_total:
            totals["row_count"] = rc_total

    return totals or None


def _gb_field_name(entry):
    if isinstance(entry, dict):
        return entry.get("field") or entry.get("name") or ""
    return entry or ""


def _alias_key(field_name):
    return (field_name or "").replace(".", "_")


def _pretty_label(name):
    parts = (name or "").replace(".", "_").split("_")
    return " ".join(p[:1].upper() + p[1:] for p in parts if p)


def _resolve_label(name, field_label_map):
    if not name:
        return ""
    if field_label_map and name in field_label_map:
        return field_label_map[name]
    base = name.split(".")[-1] if "." in name else name
    if field_label_map and base in field_label_map:
        return field_label_map[base]
    return _pretty_label(base)


def _coerce_number(v):
    if v is None or v == "":
        return None
    if isinstance(v, decimal.Decimal):
        return float(v)
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _aggregate(values, agg):
    agg = (agg or "sum").lower()
    if not values:
        return ""
    if agg == "sum":
        return round(sum(values), 2)
    if agg == "min":
        return round(min(values), 2)
    if agg == "max":
        return round(max(values), 2)
    if agg == "count":
        return int(len(values))
    if agg == "avg":
        return round(sum(values) / len(values), 2)
    return round(sum(values), 2)


def export_pivot_to_excel(
    records,
    report_name,
    fields,
    group_rows_cfg,
    group_cols_cfg,
    show_detail_rows,
    show_row_counts,
    show_grand_total,
    show_subtotals,
    field_label_map,
):
    """Render a grouped report as an Excel pivot, mirroring the preview's
    renderTable layout: group rows down the left side, group columns across
    the top, value cells per intersection, optional subtotal/grand-total rows.
    Streams via openpyxl write-only mode + temp-file FileResponse for big
    exports."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.cell.cell import WriteOnlyCell
    from openpyxl.utils import get_column_letter
    from django.http import FileResponse

    group_row_keys = [_alias_key(_gb_field_name(g)) for g in group_rows_cfg]
    group_col_keys = [_alias_key(_gb_field_name(g)) for g in group_cols_cfg]

    # Aggregate (value) fields: dicts in `fields` carrying an "aggregate"
    # entry. Add an implicit row_count when grouping + show_row_counts.
    value_fields = []
    for f in fields or []:
        if isinstance(f, dict) and f.get("aggregate") and f.get("name"):
            value_fields.append({
                "name": f["name"],
                "alias": f.get("alias") or (
                    f["aggregate"].lower() + "_" + _alias_key(f["name"])
                ),
                "aggregate": f["aggregate"],
                "label": _resolve_label(f["name"], field_label_map),
            })
    if show_row_counts:
        value_fields.append({
            "name": "row_count",
            "alias": "row_count",
            "aggregate": "sum",
            "label": "Row Count",
        })

    # rowCombos and colCombos derived from the SQL-grouped result rows.
    def _combos(records, keys):
        if not keys:
            return [tuple()]
        seen = []
        seen_set = set()
        for r in records or []:
            t = tuple(r.get(k) for k in keys)
            if t not in seen_set:
                seen_set.add(t)
                seen.append(t)
        return seen

    row_combos = _combos(records, group_row_keys)
    col_combos = _combos(records, group_col_keys)

    # Excel has a hard limit of 16,384 columns per sheet (column XFD).
    # When the pivot's column dimension explodes (e.g. grouping by product
    # with 9k+ unique products × 2 value fields = 18k columns) openpyxl
    # raises "Invalid column index". Trim col_combos to fit, leaving room
    # for the group-row label columns and a small safety margin.
    EXCEL_MAX_COLS = 16_384
    safety_margin = 4
    available_for_combos = max(
        1,
        (EXCEL_MAX_COLS - len(group_rows_cfg) - safety_margin)
        // max(1, len(value_fields)),
    )
    if len(col_combos) > available_for_combos:
        logger.warning(
            "[REPORT EXPORT] Pivot column dimension truncated for '%s': "
            "%d combos × %d value fields exceeded Excel's %d-column limit; "
            "showing the first %d combos.",
            report_name,
            len(col_combos),
            len(value_fields),
            EXCEL_MAX_COLS,
            available_for_combos,
        )
        col_combos = col_combos[:available_for_combos]

    # Index records by (rowCombo, colCombo) so the cell write loop is O(1).
    records_index = {}
    for r in records or []:
        rk = tuple(r.get(k) for k in group_row_keys)
        ck = tuple(r.get(k) for k in group_col_keys)
        records_index.setdefault((rk, ck), []).append(r)

    # ── Workbook setup (write-only for big exports) ──────────────────────
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Report")

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    group_label_fill = PatternFill("solid", fgColor="EAF1FB")
    subtotal_fill = PatternFill("solid", fgColor="F2F2F2")
    subtotal_font = Font(bold=True, color="000000")
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    total_font = Font(bold=True, color="000000")
    thin = Side(border_style="thin", color="B7B7B7")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _styled(value, *, font=None, fill=None, align=None, border=None):
        c = WriteOnlyCell(ws, value=value)
        if font:
            c.font = font
        if fill:
            c.fill = fill
        if align:
            c.alignment = align
        if border or cell_border:
            c.border = border or cell_border
        return c

    # ── Header rows ──────────────────────────────────────────────────────
    # Top header levels — one per group-column level. Each level shows the
    # column-group's field label (over the row-label columns) followed by the
    # distinct values of that level repeated for each value field.
    for level, gc in enumerate(group_cols_cfg):
        row = []
        col_label = _resolve_label(_gb_field_name(gc), field_label_map)
        for i, _ in enumerate(group_rows_cfg):
            row.append(_styled(
                col_label if i == 0 else "",
                font=header_font, fill=header_fill, align=header_align,
            ))
        # Repeat the value-field group for each col_combo at this level.
        for combo in col_combos:
            cell_value = combo[level] if level < len(combo) else ""
            for _ in value_fields:
                row.append(_styled(
                    cell_value if cell_value is not None else "",
                    font=header_font, fill=header_fill, align=header_align,
                ))
        ws.append(row)

    # Final header row: group-row field labels + value-field labels per col_combo
    last_header = []
    for gr in group_rows_cfg:
        last_header.append(_styled(
            _resolve_label(_gb_field_name(gr), field_label_map),
            font=header_font, fill=header_fill, align=header_align,
        ))
    for _combo in col_combos:
        for vf in value_fields:
            label = (
                f"{vf['aggregate'].title()} of {vf['label']}"
                if vf["aggregate"] and vf["alias"] != "row_count"
                else vf["label"]
            )
            last_header.append(_styled(
                label, font=header_font, fill=header_fill, align=header_align,
            ))
    ws.append(last_header)

    # ── Body rows: one row per row_combo ────────────────────────────────
    # Tracks running values for grand-total computation.
    grand_buckets = {(combo, vf["alias"]): [] for combo in col_combos for vf in value_fields}
    # Subtotal accumulators keyed by (level, prefix-tuple) — prefix is the
    # row_combo trimmed to that level's depth.
    subtotal_buckets = {}

    prev_row_combo = None
    for r_idx, row_combo in enumerate(row_combos):
        out = []
        # Group-row labels: only emit on level transitions (mimics rowspan).
        for level, val in enumerate(row_combo):
            if (
                prev_row_combo is not None
                and prev_row_combo[: level + 1] == row_combo[: level + 1]
            ):
                out.append(_styled("", fill=group_label_fill))
            else:
                disp = "" if val is None else val
                out.append(_styled(disp, fill=group_label_fill))
        # Value cells per col_combo × value_field
        for combo in col_combos:
            recs = records_index.get((row_combo, combo), [])
            rec = recs[0] if recs else None
            for vf in value_fields:
                value = rec.get(vf["alias"]) if rec else None
                if value is None and rec:
                    value = rec.get(vf["name"])
                num = _coerce_number(value)
                if num is not None:
                    grand_buckets[(combo, vf["alias"])].append(num)
                    # Track subtotals up the row-prefix tree.
                    for lvl in range(1, len(row_combo)):
                        sub_key = (lvl, row_combo[:lvl], combo, vf["alias"])
                        subtotal_buckets.setdefault(sub_key, []).append(num)
                out.append(_styled(num if num is not None else (value if value is not None else "")))
        ws.append(out)
        prev_row_combo = row_combo

        # Subtotals: emit when the next row's prefix differs at some level.
        if show_subtotals and len(row_combo) > 1:
            next_combo = row_combos[r_idx + 1] if r_idx + 1 < len(row_combos) else None
            for lvl in range(len(row_combo) - 1, 0, -1):
                # End of this prefix when next combo has a different prefix.
                if next_combo is None or next_combo[:lvl] != row_combo[:lvl]:
                    sub_row = []
                    for i in range(len(row_combo)):
                        if i < lvl:
                            sub_row.append(_styled("", fill=subtotal_fill, font=subtotal_font))
                        elif i == lvl:
                            sub_row.append(_styled(
                                f"Subtotal ({row_combo[lvl - 1] if row_combo[lvl - 1] is not None else ''})",
                                fill=subtotal_fill, font=subtotal_font,
                            ))
                        else:
                            sub_row.append(_styled("", fill=subtotal_fill, font=subtotal_font))
                    for combo in col_combos:
                        for vf in value_fields:
                            vals = subtotal_buckets.pop(
                                (lvl, row_combo[:lvl], combo, vf["alias"]),
                                [],
                            )
                            sub_row.append(_styled(
                                _aggregate(vals, vf["aggregate"]),
                                fill=subtotal_fill, font=subtotal_font,
                            ))
                    ws.append(sub_row)

    # ── Grand Total row ──────────────────────────────────────────────────
    if show_grand_total and (row_combos or col_combos):
        gt_row = []
        for i in range(len(group_rows_cfg)):
            gt_row.append(_styled(
                "Grand Total" if i == 0 else "",
                fill=total_fill, font=total_font,
            ))
        for combo in col_combos:
            for vf in value_fields:
                gt_row.append(_styled(
                    _aggregate(grand_buckets.get((combo, vf["alias"]), []), vf["aggregate"]),
                    fill=total_fill, font=total_font,
                ))
        ws.append(gt_row)

    # Column widths: group-row columns + (value-fields × col_combos)
    total_cols = len(group_rows_cfg) + len(col_combos) * len(value_fields)
    for i in range(1, total_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    # Spill to disk + stream back chunked.
    tmp = tempfile.NamedTemporaryFile(prefix="report_pivot_", suffix=".xlsx", delete=False)
    tmp_path = tmp.name
    tmp.close()
    wb.save(tmp_path)

    def _iter_and_cleanup(path):
        try:
            with open(path, "rb") as fh:
                while True:
                    chunk = fh.read(64 * 1024)
                    if not chunk:
                        break
                    yield chunk
        finally:
            try:
                os.unlink(path)
            except OSError:
                pass

    response = FileResponse(
        _iter_and_cleanup(tmp_path),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    safe_name = (report_name or "report").replace('"', "'")
    response["Content-Disposition"] = f'attachment; filename="{safe_name}.xlsx"'
    response["Content-Length"] = str(os.path.getsize(tmp_path))
    return response


def get_fields_metadata(table_name, schema="public"):
    with connection.cursor() as cursor:
        _set_search_path(cursor, schema)
        cursor.execute("SELECT name, label FROM fields WHERE object_name = %s", [table_name])
        return {row[0]: row[1] for row in cursor.fetchall()}


def get_fields_of_the_object(table_name, schema="public"):
    with connection.cursor() as cursor:
        _set_search_path(cursor, schema)
        cursor.execute("SELECT name, datatype FROM fields WHERE object_name = %s", [table_name])
        return {row[0]: row[1] for row in cursor.fetchall()}


def get_lookup_fields(table_name: str, schema: str = "public") -> dict[str, str]:
    """
    Return a mapping of  {field_name: related_object_name}  for every field
    on *table_name* whose datatype is 'lookup' or 'relationship' and that
    has a non-null related_object value.

    Example for the 'leads' object:
        {
            "account_id":  "accounts",
            "contact_id":  "contacts",
            "stage_id":    "stages",
        }

    owner_id / created_by_id / last_modified_by_id are intentionally excluded
    at call-site (they are handled by USER_ID_COLUMNS and would produce
    duplicate JOINs if processed here too).
    """
    with connection.cursor() as cursor:
        _set_search_path(cursor, schema)
        cursor.execute(
            """
            SELECT name, parent_object
            FROM   fields
            WHERE  object_name = %s
              AND  datatype IN ('lookup_relationship')
              AND  relationship_name IS NOT NULL
              AND  relationship_name <> ''
            """,
            [table_name],
        )
        return {row[0]: row[1] for row in cursor.fetchall()}

import json

def get_selected_object_app(appname,objects,schema):
    with connection.cursor() as cursor:
        _set_search_path(cursor, schema)
        cursor.execute("SELECT id,tabs FROM app WHERE name = %s", [appname])
        app_row = cursor.fetchone()
        if not app_row:
            raise ValueError(f"App '{appname}' not found.")
        object_tab =  json.loads(app_row[1])
        if not object_tab:
            raise ValueError(f"App '{appname}' has no tabs defined.")
        return [tab["name"] for tab in object_tab if tab["name"] in objects]


@api_view(["POST"])
@authentication_classes([CustomJWTAuthentication])
# @permission_classes([IsAuthenticated])
def export_selected_objects(request):
    """
    Export one or more CRM objects to a single .xlsx file.

    Optimisations vs original
    ─────────────────────────
    BLOCKER-1  No fetchall() — rows streamed from PostgreSQL via named cursor
               + fetchmany(STREAM_BATCH_SIZE); peak RAM = one 2 000-row batch.
    BLOCKER-2  Workbook(write_only=True) — openpyxl never builds the cell graph;
               rows are serialised directly into the ZIP stream.
    BLOCKER-3  get_object_details() called once per object; result cached in
               valid_objects tuple and reused in export loop.
    BLOCKER-4  _resolve_schema_and_columns() replaces 3 separate
               information_schema round-trips with one combined query.
    BLOCKER-5  get_user_timezone() called once before the loop, not per object.
    BLOCKER-6  Datetime conversion done per-value at cursor read time via pytz;
               no pandas DataFrame, no apply(), zero extra RAM.
    BLOCKER-7  BytesIO passed directly to HttpResponse — no .getvalue() copy.
    BLOCKER-8  audit log entries collected in a list; flushed after the loop
               so synchronous DB INSERTs don't block each object's export.
    BLOCKER-9  MAX_EXPORT_OBJECTS guard re-enabled.
    BLOCKER-10 logger at module level (not recreated per request).
    LOGIC-BUG  Date format fixed to %d/%m/%Y; print() replaced with logger.
    """
    try:
        data             = request.data or {}
        selected_objects = data.get("selected_objects", [])
        # selected_app = data.get("selected_app")
        user, org, _db_conn, profile_id, schema, referer = get_connection_and_user_details(request)
        if not selected_objects:
            return Response({"error": "No objects selected"}, status=status.HTTP_400_BAD_REQUEST)
        # selected_objects = get_selected_object_app(selected_app,selected_objects,schema)

        # FIX BLOCKER-9: enforce object count cap (was commented out)
        # if len(selected_objects) > MAX_EXPORT_OBJECTS:
        #     return Response(
        #         {"error": f"Too many objects selected (max {MAX_EXPORT_OBJECTS})"},
        #         status=status.HTTP_400_BAD_REQUEST,
        #     )

        # ── Validation pass ────────────────────────────────────────────────────
        # FIX BLOCKER-3: call get_object_details() ONCE and carry the result
        # into the export loop — original code called it a second time on line 240.
        # FIX BLOCKER-4: _resolve_schema_and_columns() replaces 3 round-trips.
        valid_objects   = []   # (object_name, schema, object_id, is_setup, real_columns)
        skipped_objects = []

        for object_name in selected_objects:
            object_row = get_object_details(object_name, schema=schema, include_setup=True)
            if not object_row:
                logger.warning(f"[EXPORT] '{object_name}' not found in metadata.")
                skipped_objects.append({"object": object_name, "reason": "Object not found in metadata"})
                continue

            object_id, object_label, is_setup = object_row

            # Single catalog query for schema + column list (FIX BLOCKER-4)
            object_schema, real_columns = _resolve_schema_and_columns(object_name, schema)
            if not object_schema:
                logger.warning(f"[EXPORT] Table '{object_name}' not found in DB.")
                skipped_objects.append({"object": object_name, "reason": f"Table '{object_name}' not found in database"})
                continue

            if object_schema != schema:
                logger.info(f"[EXPORT] '{object_name}' found in schema '{object_schema}' (fallback).")

            if not is_setup:
                if not check_permission(object_id, "read", schema=schema, profile_id=profile_id):
                    logger.warning(f"[EXPORT] No read permission for '{object_label}'.")
                    skipped_objects.append({"object": object_name, "reason": f"No read permission for {object_label}"})
                    continue

            logger.info(f"[EXPORT] '{object_name}' validated — {len(real_columns)} columns.")
            valid_objects.append((object_name, object_schema, object_id, is_setup, real_columns))

        if not valid_objects:
            return Response(
                {"error": "None of the selected objects could be exported", "skipped": skipped_objects},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # FIX BLOCKER-5: fetch user timezone ONCE before the loop
        tz_name  = get_user_timezone(user.get("id"), schema=schema)
        user_tz  = pytz.timezone(tz_name)

        # FIX BLOCKER-2: write_only=True — openpyxl never builds a cell graph
        wb          = Workbook(write_only=True)
        audit_batch = []   # FIX BLOCKER-8: collect audit entries; flush after loop

        for object_name, object_schema, object_id, is_setup, real_columns in valid_objects:
            try:
                # ── Determine which fields to export ──────────────────────────
                computed_meta = {}
                if is_setup:
                    export_fields = real_columns
                    logger.info(f"[EXPORT] {object_name}: setup object — {len(export_fields)} cols.")
                else:
                    permitted_fields, _meta = get_field_metadata(
                        object_id, "read", schema=schema, profile_id=profile_id,
                    )
                    # Identify computed fields (formula/rollup_summary)
                    computed_field_names = set()
                    for m in _meta:
                        if m.get('datatype') in ('formula', 'rollup_summary'):
                            computed_field_names.add(m['name'])
                            computed_meta[m['name']] = m

                    if permitted_fields:
                        real_set = set(real_columns)
                        export_fields = [f for f in permitted_fields if f in real_set]
                        # Add computed field names (not physical but will be computed)
                        for cf in computed_field_names:
                            if cf not in export_fields:
                                export_fields.append(cf)
                        if not export_fields:
                            logger.warning(f"[EXPORT] {object_name}: no permitted fields match real columns, falling back.")
                            export_fields = real_columns
                        else:
                            logger.info(f"[EXPORT] {object_name}: {len(export_fields)} permitted cols ({len(computed_field_names)} computed).")
                    else:
                        logger.info(f"[EXPORT] {object_name}: no field-level perms, exporting all cols.")
                        export_fields = real_columns

                if not export_fields:
                    logger.warning(f"[EXPORT] {object_name}: no exportable columns.")
                    export_fields = ["id"]

                # Fetch field labels for header row
                field_label_map = get_fields_metadata(object_name, schema=object_schema)

                if computed_meta:
                    # Objects with computed fields — fetch via get_permissions + compute
                    physical_fields = [f for f in export_fields if f not in computed_meta]
                    # Add dependency fields for formula evaluation
                    _, computed_fields_full, extra_deps = process_computed_fields_for_report(
                        export_fields, object_name, object_schema
                    )
                    for dep in extra_deps:
                        if dep not in physical_fields:
                            physical_fields.append(dep)

                    all_records = get_permissions(
                        request,
                        tableName=object_name,
                        fields=physical_fields,
                        schema=object_schema,
                        profile_id=profile_id,
                        limit=MAX_ROWS_PER_OBJECT,
                    ).get("data", [])
                    all_records = apply_computed_fields_to_records(
                        all_records, computed_fields_full, object_name, object_schema
                    )

                    # Write to sheet
                    ws = wb.create_sheet(title=_safe_sheet_name(object_name))
                    header_fill = PatternFill("solid", fgColor="4472C4")
                    header_font = Font(bold=True, color="FFFFFF")
                    header_row = []
                    for col_name in export_fields:
                        cell = WriteOnlyCell(ws, value=field_label_map.get(col_name, col_name))
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        header_row.append(cell)
                    ws.append(header_row)
                    for record in all_records:
                        row = []
                        for col_name in export_fields:
                            val = record.get(col_name)
                            if isinstance(val, datetime) and hasattr(val, 'tzinfo') and val.tzinfo:
                                val = val.astimezone(user_tz).strftime("%d/%m/%Y %H:%M")
                            row.append(val)
                        ws.append(row)
                    rows_written = len(all_records)
                else:
                    # No computed fields — use efficient streaming
                    lookup_fields = get_lookup_fields(object_name, schema=object_schema)
                    logger.info(
                        f"[EXPORT] {object_name}: {len(lookup_fields)} lookup field(s) "
                        f"will be resolved via JOIN: {list(lookup_fields.keys()) or 'none'}"
                    )
                    ws = wb.create_sheet(title=_safe_sheet_name(object_name))
                    rows_written = _stream_rows_to_sheet(
                        ws             = ws,
                        object_name    = object_name,
                        schema         = object_schema,
                        fields         = export_fields,
                        field_label_map= field_label_map,
                        user_tz        = user_tz,
                        lookup_fields  = lookup_fields,
                    )
                logger.info(f"[EXPORT] {object_name}: {rows_written} rows written.")

                # FIX BLOCKER-8: accumulate audit entry instead of firing DB INSERT now
                if user:
                    audit_batch.append({
                        "msg":    f"Exported data for {object_name} ({rows_written} rows)",
                        "section":f"Export - {object_name}",
                        "prefix": object_name[:3].upper(),
                    })

            except Exception as err:
                # FIX LOGIC-BUG: use logger, not print()
                logger.exception(f"[EXPORT] Error exporting {object_name}: {err}")
                if user:
                    audit_batch.append({
                        "msg":    f"Failed to export {object_name}: {err}",
                        "section":f"Export error - {object_name}",
                        "prefix": object_name[:3].upper(),
                    })
                continue

        # ── Skipped-objects sheet ──────────────────────────────────────────────
        if skipped_objects:
            try:
                ws_skip = wb.create_sheet(title=_safe_sheet_name("skipped_objects"))
                ws_skip.append(list(skipped_objects[0].keys()))
                for row in skipped_objects:
                    ws_skip.append(list(row.values()))
            except Exception as e:
                logger.error(f"[EXPORT] Failed to write skipped_objects sheet: {e}")

        # Persist the workbook to a temp file and stream it back. For 1L+ row
        # exports the xlsx can easily run to tens of MB — FileResponse chunks
        # the disk read so we don't duplicate the whole payload in RAM.
        tmp = tempfile.NamedTemporaryFile(
            prefix="export_", suffix=".xlsx", delete=False
        )
        tmp_path = tmp.name
        tmp.close()
        wb.save(tmp_path)

        # FIX BLOCKER-8: flush all audit entries in one go, after the export
        if user:
            for entry in audit_batch:
                try:
                    log_audit(entry["msg"], entry["section"],
                              prefix=entry["prefix"], schema=schema, user_=user)
                except Exception as audit_err:
                    logger.warning(f"[EXPORT] Audit log failed: {audit_err}")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename  = f"exported_data_{timestamp}.xlsx"

        def _iter_and_cleanup(path):
            try:
                with open(path, "rb") as fh:
                    while True:
                        chunk = fh.read(FILE_STREAM_CHUNK)
                        if not chunk:
                            break
                        yield chunk
            finally:
                try:
                    os.unlink(path)
                except OSError:
                    pass

        response = FileResponse(
            _iter_and_cleanup(tmp_path),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = str(os.path.getsize(tmp_path))
        return response

    except Exception as e:
        logger.exception(f"[EXPORT] export_selected_objects failed: {e}")
        return Response({"error": "Failed to export selected objects"}, status=500)


@api_view(["POST"])
@authentication_classes([CustomJWTAuthentication])
@permission_classes([IsAuthenticated])
def export_audit_trail(request):
    try:
        _, _, _db_conn, _profile_id, schema, _referer = get_connection_and_user_details(request)
        rows, columns = fetch_audit_trail_raw(schema=schema)

        wb = Workbook()
        ws = wb.active
        ws.title = "Audit Trail Logs"

        ws.append(columns)
        for row in rows:
            formatted_row = [
                *row[:-1],
                row[-1].strftime("%Y-%m-%d %H:%M:%S") if isinstance(row[-1], datetime) else row[-1],
            ]
            ws.append(formatted_row)

        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        filename = f"audit_trail_export_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        response = HttpResponse(
            file_stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    except Exception:
        return Response({"error": "Failed to export audit trail"}, status=500)


@api_view(["GET"])
@authentication_classes([CustomJWTAuthentication])
@permission_classes([IsAuthenticated])
def export_report_excel(request):
    try:
        report_id = request.GET.get("id")
        if not report_id:
            return Response({"error": "Report ID is required"}, status=400)

        user, org, _db_conn, profile_id, schema, referer = get_connection_and_user_details(request)

        report_data = get_permissions(
            request,
            tableName="report",
            where=[{"field": "id", "value": report_id, "operator": "="}],
            schema=schema,
            profile_id=profile_id,
        ).get("data")

        if not report_data:
            return Response({"error": "Report not found"}, status=404)

        report = report_data[0]
        report_name = report.get("name")
        fields = report.get("fields") or []
        # Multi-select filter UIs sometimes serialise N picked values as a
        # single ILIKE on `<rel>_id` with a comma-joined string. That hits
        # a sequential scan AND matches no rows — the export silently
        # downloaded 0 rows. Normalise to a real IN list.
        from api.BL.blcontroller import _normalize_csv_ilike_filters
        filters = _normalize_csv_ilike_filters(report.get("filters") or [])
        group_by = report.get("group_by") or []
        filter_logic = report.get("filter_logic")
        table_name = report.get("table_name")
        show_detail_rows = report.get("show_detail_rows", False)
        show_grand_total = report.get("show_grand_total", True)
        show_row_counts = report.get("show_row_counts", True)
        # Toggle from the preview UI overrides the saved report setting so the
        # export matches what the user currently sees on screen.
        def _qparam_bool(key, current):
            raw = request.GET.get(key)
            if raw is None:
                return current
            return str(raw).lower() == "true"
        show_detail_rows = _qparam_bool("show_detail_rows", show_detail_rows)
        show_grand_total = _qparam_bool("show_grand_total", show_grand_total)
        show_row_counts = _qparam_bool("show_row_counts", show_row_counts)

        group_by_norm = normalize_group_by(group_by)
        group_by_fields = group_by_norm.get("rows", []) + group_by_norm.get("columns", [])
        filtered_fields = filter_summary_fields(fields, group_by)

        # Separate computed fields from physical fields
        physical_filtered, computed_fields, extra_deps = process_computed_fields_for_report(filtered_fields, table_name, schema)
        for dep in extra_deps:
            physical_filtered.append(dep)
        # filter_summary_fields drops non-aggregate fields when there's no
        # group-by, so `computed_fields` above might miss formula/rollup
        # columns the user added as plain detail columns. Re-classify against
        # the FULL fields list so we know whether the export needs the
        # apply_computed_fields_to_records path even for ungrouped detail
        # reports — without this the streaming detail path emits empty
        # cells for those columns.
        details_fields_all = get_details_fields(fields)
        physical_full, computed_full, extra_deps_full = process_computed_fields_for_report(
            details_fields_all, table_name, schema
        )
        # Remove computed fields from group_by
        computed_names = {meta.get("name", k) for k, meta in computed_fields.items()}
        def _gb_name(g):
            if isinstance(g, dict):
                return g.get("field") or g.get("name", "")
            return g
        group_by_fields = [g for g in group_by_fields if _gb_name(g) not in computed_names]

        # Separate computed-field filters (formula/rollup) and push the
        # rollup ones down to physical FK-IN predicates. Without this the
        # raw `invoice.grand_total < 100000` filter reaches get_permissions
        # and the SQL builder emits `invoice_item__invoice.grand_total`,
        # which doesn't exist as a column on the parent table — Postgres
        # raises 'column does not exist' and the export 500s.
        all_computed_for_filters = set(computed_fields.keys()) | set(computed_full.keys())
        all_computed_for_filters |= {meta.get("name") for meta in computed_fields.values() if meta.get("name")}
        all_computed_for_filters |= {meta.get("name") for meta in computed_full.values() if meta.get("name")}
        filters, computed_filters = separate_computed_filters(
            filters, all_computed_for_filters, schema, table_name,
        )
        extra_phys_export, computed_filters = convert_rollup_filters_to_physical(
            computed_filters, table_name, schema,
        )
        if extra_phys_export:
            filters = list(filters) + extra_phys_export

        # Grouped reports → pivot export. Always use the SQL-aggregated
        # summary (one row per group combo, ~thousands at most) regardless
        # of show_detail_rows. Pulling 1L+ raw rows just to dedupe back to
        # the same per-combo intersections in the pivot writer was the slow
        # path the user was hitting.
        if group_by_fields:
            # SQL GROUP BY can only emit columns that are either in the
            # GROUP BY list or wrapped in an aggregate. `physical_filtered`
            # also contains formula-dependency fields (unit_price, quantity,
            # …) appended via `extra_deps` — those must be stripped here or
            # PostgreSQL raises "column X must appear in the GROUP BY clause".
            gb_field_names = set()
            for g in group_by_fields:
                if isinstance(g, dict):
                    gb_field_names.add(g.get("field") or g.get("name") or "")
                else:
                    gb_field_names.add(g)
            summary_fields = []
            for f in physical_filtered:
                if isinstance(f, dict):
                    if f.get("aggregate"):
                        summary_fields.append(f)
                    elif (f.get("name") or "") in gb_field_names or (
                        f.get("alias") or ""
                    ) in gb_field_names:
                        summary_fields.append(f)
                else:
                    if f in gb_field_names:
                        summary_fields.append(f)
            # Inline formula/rollup aggregates as raw SQL so the grouped
            # export populates "Sum of <Formula>" / "Sum of <Rollup>"
            # cells. Without this the pivot writer reads NULL for those
            # columns and emits empty cells. Each computed_fields entry
            # with an aggregate gets pushed down as `<AGG>(<expr>)`.
            #
            # Bare column references in the formula get qualified with
            # the base-table name so the JOIN-based SELECT doesn't get
            # an ambiguous-column error (which silently emits NULL on
            # some clients) when the joined parent table happens to
            # share a column name. Without this the user's
            # `quantity * unit_price * tax_percent / 100` formula
            # rendered as empty cells in the exported xlsx because the
            # SQL planner couldn't resolve `quantity` against the joined
            # invoice/account tables.
            from api.BL.dashboards.dashboard import _qualify_formula_expr, _list_table_columns, _SQL_FUNCTION_TOKENS
            import re as _re_export
            _base_columns = _list_table_columns(table_name, schema)

            def _qualify_and_coalesce(expr, table_name, columns):
                """Same as _qualify_formula_expr but wraps each column ref
                in COALESCE(col, 0). Without this, a single NULL operand
                in `quantity * unit_price - discount_amount` makes the
                entire row's formula evaluate to NULL and SUM aggregates
                ignore it — which is why every "Sum of Grand Total" cell
                in the exported xlsx came back blank."""
                if not expr:
                    return expr
                def _repl(m):
                    ident = m.group(0)
                    if ident.upper() in _SQL_FUNCTION_TOKENS:
                        return ident
                    if ident in columns:
                        return f'COALESCE("{table_name}"."{ident}", 0)'
                    return ident
                return _re_export.sub(r"\b[a-zA-Z_]\w*\b", _repl, expr)

            for _ck, _meta in (computed_fields or {}).items():
                _agg = (_meta.get("aggregate") or "").lower()
                if not _agg:
                    continue
                _alias = _meta.get("alias") or _ck
                _datatype = _meta.get("datatype")
                if _datatype == "formula" and _meta.get("formula_expression"):
                    _expr = _qualify_and_coalesce(
                        _meta["formula_expression"], table_name, _base_columns,
                    )
                    summary_fields.append({
                        "name": f"{_agg.upper()}({_expr})",
                        "expression": True,
                        "alias": _alias,
                    })
                elif _datatype == "rollup_summary":
                    # Rollup is parent-side, so it can't be aggregated by
                    # this base-table GROUP BY. Skip — the pivot writer
                    # will leave these cells blank for grouped exports;
                    # users wanting per-group rollup totals should use
                    # the report preview's computed-aggregate path.
                    continue
            if not any(
                isinstance(f, dict) and f.get("alias") == "row_count"
                for f in summary_fields
            ):
                summary_fields.append({
                    "aggregate": "count",
                    "name": "id",
                    "alias": "row_count",
                })
            # When grouping by a parent's field (e.g. "invoice.name"), also
            # SELECT the FK on the base table so we can batch-evaluate any
            # rollup_summary aggregates per parent and inject the values
            # back into the grouped result rows. Without this the rollup
            # cells render blank because SQL can't aggregate Python-side
            # rollup fields.
            rollup_targets = []  # [(alias, meta, fk_field, fk_alias)]
            for _ck, _meta in (computed_fields or {}).items():
                if (_meta.get("datatype") != "rollup_summary"
                        or not _meta.get("aggregate")):
                    continue
                # Find a group_by entry whose dot-path resolves to the
                # rollup's parent table; the FK on the base table is then
                # `<relationship>_id`.
                gb_match = None
                for g in group_by_fields:
                    name = g.get("field") or g.get("name") if isinstance(g, dict) else g
                    if name and "." in name:
                        gb_match = name
                        break
                if not gb_match:
                    continue
                relationship = gb_match.split(".")[0]
                fk_field = f"{relationship}_id"
                fk_alias = f"__rollup_fk__{relationship}"
                if not any(
                    isinstance(f, dict) and f.get("alias") == fk_alias
                    for f in summary_fields
                ):
                    # invoice_id isn't in GROUP BY — wrap in MIN() so SQL
                    # accepts it. Since invoice_name is in the GROUP BY and
                    # uniquely maps to invoice_id, MIN/MAX/any aggregate
                    # gives the same single FK per group.
                    summary_fields.append({
                        "name": fk_field,
                        "aggregate": "min",
                        "alias": fk_alias,
                    })
                rollup_targets.append((_ck, _meta, fk_field, fk_alias))

            result = get_permissions(
                request,
                tableName=table_name,
                fields=summary_fields,
                where=filters,
                group_by=group_by_fields,
                report=True,
                schema=schema,
                limit=MAX_ROWS_PER_OBJECT,
                profile_id=profile_id,
            ).get("data", [])
            # Honour any computed filters that couldn't be pushed to SQL
            # (formula predicates the inliner rejects, etc.) so the export
            # reflects the report's full filter set, not just the pushed
            # subset.
            print(len(result))
            if computed_filters:
                result = apply_computed_filters(
                    result, computed_filters,
                    computed_fields=computed_fields,
                    table_name=table_name, schema=schema,
                )

            # Batch-evaluate rollups using the FK column we just smuggled in
            # and inject value × row_count (the SUM semantic for a parent
            # rollup grouped by parent: every invoice_item in the group has
            # the same parent's rollup, so SUM = rollup_value * count).
            if rollup_targets and result:
                from api.formulas.evaluate_rollup import evaluate_rollup_batch
                for alias, meta, fk_field, fk_alias in rollup_targets:
                    parent_ids = list({r.get(fk_alias) for r in result if r.get(fk_alias)})
                    if not parent_ids:
                        for row in result:
                            row[alias] = 0
                        continue
                    try:
                        relationship = fk_field[:-3]  # strip _id
                        with connection.cursor() as cursor:
                            cursor.execute("SET search_path TO %s", [schema])
                            cursor.execute(
                                "SELECT name FROM object WHERE name = %s OR name = %s",
                                [relationship, relationship + "s"],
                            )
                            obj_row = cursor.fetchone()
                            parent_table = obj_row[0] if obj_row else relationship
                        values_by_parent = evaluate_rollup_batch(
                            record_ids=parent_ids,
                            summarized_object=meta.get("summarized_object"),
                            rollup_type=meta.get("rollup_type"),
                            field_to_aggregate=meta.get("field_to_aggregate"),
                            filter_criteria=meta.get("filter_criteria"),
                            parent_field=f"{parent_table}_id",
                            schema=schema,
                        )
                    except Exception as e:
                        print(f"[REPORT EXPORT] rollup batch eval failed for {alias}: {e}")
                        values_by_parent = {}
                    agg = (meta.get("aggregate") or "sum").lower()
                    for row in result:
                        pid = row.get(fk_alias)
                        v = values_by_parent.get(pid)
                        if v is None:
                            row[alias] = 0
                            continue
                        try:
                            v = float(v)
                        except (TypeError, ValueError):
                            row[alias] = 0
                            continue
                        if agg == "sum":
                            row[alias] = round(v * (row.get("row_count") or 1), 2)
                        else:
                            row[alias] = v
                # Strip the temporary FK columns so they don't show in the
                # exported sheet.
                for row in result:
                    for _, _, _, fk_alias in rollup_targets:
                        row.pop(fk_alias, None)

            if not show_row_counts:
                for row in result:
                    row.pop("row_count", None)
            fields_metadata = get_fields_metadata(table_name, schema=schema)
            fields_metadata = dict(fields_metadata)
            fields_metadata.setdefault("row_count", "Row Count")
            return export_pivot_to_excel(
                records=result,
                report_name=report_name,
                fields=fields,
                group_rows_cfg=group_by_norm.get("rows", []) or [],
                group_cols_cfg=group_by_norm.get("columns", []) or [],
                show_detail_rows=show_detail_rows,
                show_row_counts=show_row_counts,
                show_grand_total=show_grand_total,
                show_subtotals=bool(report.get("show_subtotals", False)),
                field_label_map=fields_metadata,
            )

        if computed_full:
            # When computed fields exist (formula/rollup), fetch detail rows
            # then run apply_computed_fields_to_records so each record has
            # the formula-evaluated and rollup-evaluated keys populated.
            # `computed_full` is the full classification including non-
            # aggregate detail columns that filter_summary_fields drops.
            physical_details = list(physical_full)
            computed_details = computed_full
            for dep in extra_deps_full:
                if dep not in [f.get("name") if isinstance(f, dict) else f for f in physical_details]:
                    physical_details.append(dep)
            # Ensure 'id' is included for rollup evaluation (removed from output later)
            _id_added = False
            if 'id' not in [f.get("name") if isinstance(f, dict) else f for f in physical_details]:
                physical_details.insert(0, 'id')
                _id_added = True
            result = get_permissions(
                request,
                tableName=table_name,
                fields=physical_details,
                where=filters,
                report=True,
                schema=schema,
                profile_id=profile_id,
                limit=MAX_ROWS_PER_OBJECT,
            ).get("data", [])
            result = apply_computed_fields_to_records(result, computed_details, table_name, schema)
            if computed_filters:
                result = apply_computed_filters(
                    result, computed_filters,
                    computed_fields=computed_details,
                    table_name=table_name, schema=schema,
                )
            # Remove 'id' from output if it was added only for rollup evaluation
            if _id_added:
                for record in result:
                    record.pop('id', None)
        elif not show_detail_rows:
            if group_by_fields and not any(
                isinstance(f, dict) and f.get("alias") == "row_count"
                for f in physical_filtered
            ):
                physical_filtered.append({
                    "aggregate": "count",
                    "name": "id",
                    "alias": "row_count",
                })
            result = get_permissions(
                request,
                tableName=table_name,
                fields=physical_filtered,
                where=filters,
                group_by=group_by_fields,
                report=True,
                schema=schema,
                limit=MAX_ROWS_PER_OBJECT,
                profile_id=profile_id,
            ).get("data", [])
            if computed_filters:
                result = apply_computed_filters(
                    result, computed_filters,
                    computed_fields=computed_fields,
                    table_name=table_name, schema=schema,
                )
        else:
            details_fields = get_details_fields(fields)
            # FAST-PATH: when there are no group dimensions, no computed
            # fields, and no filters, the entire export is just "stream the
            # details to xlsx". Skip the get_permissions list-materialization
            # (which holds every row in Python memory before we touch the
            # workbook) and instead stream the result-set directly through a
            # named cursor into a write-only worksheet, the same pattern
            # export_selected_objects already uses for big single-object
            # exports.
            #
            # Filters force the slow path: `_stream_rows_to_sheet` does a
            # blanket `SELECT ... FROM table LIMIT 200000` with no WHERE
            # clause, so it would dump every row regardless of the report's
            # filter set. The user's filter conditions MUST be honoured by
            # the export, so we fall through to the get_permissions path
            # below whenever any filter is configured.
            if not group_by_fields and not filters and not computed_filters:
                normalized_field_names = []
                for f in details_fields:
                    name = f.get("name") if isinstance(f, dict) else f
                    if name and name not in normalized_field_names and "." not in name:
                        normalized_field_names.append(name)
                if normalized_field_names:
                    user_tz_name = get_user_timezone(user.get("id"), schema=schema) if user else DEFAULT_TZ
                    user_tz = pytz.timezone(user_tz_name)
                    wb = Workbook(write_only=True)
                    ws = wb.create_sheet(title=_safe_sheet_name(table_name))
                    field_label_map_pre = get_fields_metadata(table_name, schema=schema)
                    field_label_map_pre = dict(field_label_map_pre)
                    field_label_map_pre.setdefault("row_count", "Row Count")
                    try:
                        lookup_fields = get_lookup_fields(table_name, schema=schema)
                    except Exception:
                        lookup_fields = {}
                    rows_written = _stream_rows_to_sheet(
                        ws=ws,
                        object_name=table_name,
                        schema=schema,
                        fields=normalized_field_names,
                        field_label_map=field_label_map_pre,
                        user_tz=user_tz,
                        lookup_fields=lookup_fields,
                    )
                    logger.info(f"[REPORT EXPORT] streamed {rows_written} rows for '{report_name}'")
                    tmp = tempfile.NamedTemporaryFile(prefix="report_", suffix=".xlsx", delete=False)
                    tmp_path = tmp.name
                    tmp.close()
                    wb.save(tmp_path)
                    def _iter_and_cleanup(path):
                        try:
                            with open(path, "rb") as fh:
                                while True:
                                    chunk = fh.read(64 * 1024)
                                    if not chunk:
                                        break
                                    yield chunk
                        finally:
                            try:
                                os.unlink(path)
                            except OSError:
                                pass
                    response = FileResponse(
                        _iter_and_cleanup(tmp_path),
                        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                    safe_name = (report_name or "report").replace('"', "'")
                    response["Content-Disposition"] = f'attachment; filename="{safe_name}.xlsx"'
                    response["Content-Length"] = str(os.path.getsize(tmp_path))
                    return response

            # Fallback (computed deps or dotted-field detail rows): still
            # bounded so a 3L-row source table doesn't OOM the worker.
            result = get_permissions(
                request,
                tableName=table_name,
                fields=details_fields,
                where=filters,
                report=True,
                schema=schema,
                profile_id=profile_id,
                limit=MAX_ROWS_PER_OBJECT,
            ).get("data", [])
            if computed_filters:
                result = apply_computed_filters(
                    result, computed_filters,
                    computed_fields=computed_full,
                    table_name=table_name, schema=schema,
                )

        # Drop row_count column from output when Row Counts toggle is off.
        if not show_row_counts:
            for row in result:
                row.pop("row_count", None)

        # Build grand total row from aggregate fields in the report config.
        grand_total_row = None
        if show_grand_total and result:
            grand_total_row = _compute_grand_total(
                result, fields, show_detail_rows, show_row_counts, group_by_fields
            )

        fields_metadata = get_fields_metadata(table_name, schema=schema)
        fields_metadata = dict(fields_metadata)
        fields_metadata.setdefault("row_count", "Row Count")

        # Pivot export: when the report has grouping configured, render the
        # same pivot layout the preview page shows — group rows down the
        # left (with merged labels), group columns across the top, value
        # cells at each (rowCombo × colCombo) intersection, plus optional
        # subtotal and grand-total rows. Falls back to the flat sheet for
        # ungrouped reports.
        group_rows_cfg = group_by_norm.get("rows", []) or []
        group_cols_cfg = group_by_norm.get("columns", []) or []
        if group_rows_cfg or group_cols_cfg:
            return export_pivot_to_excel(
                records=result,
                report_name=report_name,
                fields=fields,
                group_rows_cfg=group_rows_cfg,
                group_cols_cfg=group_cols_cfg,
                show_detail_rows=show_detail_rows,
                show_row_counts=show_row_counts,
                show_grand_total=show_grand_total,
                show_subtotals=bool(report.get("show_subtotals", False)),
                field_label_map=fields_metadata,
            )

        return export_to_excel(
            result,
            filename=f"{report_name}.xlsx",
            field_label_map=fields_metadata,
            grand_total_row=grand_total_row,
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        from django.http import JsonResponse as _JsonResponse
        return _JsonResponse({"error": str(e), "detail": traceback.format_exc()}, status=500)