from collections import defaultdict
import json

from utils.string_converters import to_camel_case_with_spaces

def process_component_data(component_type, metric_config, chart_config, filtered_data):
    """
    Process a component based on its type and return the computed data.
    """
    component_data = {}    
    try:
        # ✅ Process Metric Components
        if component_type == "metric":
            if filtered_data:          
                if "avg" in metric_config:
                    field = metric_config["avg"]
                    values = [entry[field] for entry in filtered_data if field in entry and entry[field] is not None]
                    return sum(values) / len(values) if values else 0
                if "max" in metric_config:
                    field = metric_config["max"]
                    values = [entry[field] for entry in filtered_data if field in entry and entry[field] is not None]
                    return max(values) if values else None
                if "min" in metric_config:
                    field = metric_config["min"]
                    values = [entry[field] for entry in filtered_data if field in entry and entry[field] is not None]
                    return min(values) if values else None                    
                if "total" in metric_config:
                    field = metric_config["total"]
                    values = [entry[field] for entry in filtered_data if field in entry and entry[field] is not None]
                    return sum(values) if values else 0          
                return len(filtered_data)  # Count total records
        # ✅ Process Chart Components
        elif "chart" in component_type:           
            group_by_field = chart_config.get("group_by", "status")  # Default to status if not provided
            chart_data = defaultdict(int)
            for entry in filtered_data:
                key = entry.get(group_by_field, "Unknown")
                chart_data[key] += 1
            # Create the list of labels and values
            labels = list(chart_data.keys())  # List of unique values (labels)
            values = list(chart_data.values())  # List of counts (values)
            component_data = {
                "labels": labels,
                "values": values
            }
        # ✅ Process Table Components
        elif component_type == "table":
            visible_columns = chart_config.get("visible_columns", ["id", "name"])  # Default columns if not provided
            component_data = [{col: entry.get(col, None) for col in visible_columns} for entry in filtered_data]
        return component_data
    except Exception as e:
        return len(filtered_data)

def safe_json_parse(value):
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return {}
    return {}

def build_folder_tree(folders, **kwargs):
    folder_map = {folder['id']: {**folder, "children": []} for folder in folders}
    root_folders = []
    for folder in folders:
        parent_id = folder.get('parent_id')
        if parent_id and parent_id in folder_map:
            folder_map[parent_id]["children"].append(folder_map[folder['id']])
        else:
            root_folders.append(folder_map[folder['id']])
    return root_folders

def normalize_row_keys(row):
    new_row = {}
    for key, value in row.items():
        if isinstance(key, dict):
            new_key = key.get('alias') or key.get('name')
        else:
            new_key = key
        new_row[str(new_key)] = value
    return new_row

def export_to_excel(data, filename="report.xlsx", field_label_map=None, grand_total_row=None, grand_total_label="Grand Total"):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.cell.cell import WriteOnlyCell
        from openpyxl.utils import get_column_letter
        from django.http import FileResponse
        import decimal
        import os
        import tempfile

        # write_only mode = constant-memory: rows are streamed straight into the
        # zipped xlsx part on append() instead of materialising a cell graph in
        # RAM. Critical for 100K+ row report exports.
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="Report")

        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        total_fill = PatternFill("solid", fgColor="FFF2CC")
        total_font = Font(bold=True, color="000000")

        def _styled_header(value):
            c = WriteOnlyCell(ws, value=value)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_alignment
            return c

        def _styled_total(value):
            c = WriteOnlyCell(ws, value=value)
            c.font = total_font
            c.fill = total_fill
            return c

        data_iter = iter(data or [])
        first_row = None
        for candidate in data_iter:
            first_row = candidate
            break

        rows_written = 0
        raw_headers = []

        if first_row is None:
            ws.append([_styled_header("S.No"), _styled_header("No data found")])
        else:
            first_normalized = normalize_row_keys(first_row)
            raw_headers = list(first_normalized.keys())
            headers = (
                [field_label_map.get(h, to_camel_case_with_spaces(h)) for h in raw_headers]
                if field_label_map else raw_headers
            )
            ws.append([_styled_header(h) for h in (["S.No"] + list(headers))])

            def _emit(idx, normalized):
                row_data = [idx]
                for col in raw_headers:
                    value = normalized.get(col, "")
                    if isinstance(value, decimal.Decimal):
                        value = float(value)
                    elif value is None:
                        value = ""
                    row_data.append(value)
                ws.append(row_data)

            _emit(1, first_normalized)
            rows_written = 1
            for row in data_iter:
                rows_written += 1
                _emit(rows_written, normalize_row_keys(row))

            if grand_total_row:
                total_row_data = [_styled_total(grand_total_label)]
                label_placed = True
                for col in raw_headers:
                    if col in grand_total_row:
                        v = grand_total_row[col]
                        if isinstance(v, decimal.Decimal):
                            v = float(v)
                        total_row_data.append(_styled_total(v if v is not None else ""))
                    elif label_placed:
                        total_row_data.append(_styled_total(""))
                    else:
                        total_row_data.append(_styled_total(grand_total_label))
                        label_placed = True
                ws.append(total_row_data)

            ws.column_dimensions[get_column_letter(1)].width = 8
            for i, col in enumerate(headers, 2):
                ws.column_dimensions[get_column_letter(i)].width = max(len(str(col)), 15)

        print(f"Data written to Excel: {rows_written} rows")

        # Spill the workbook to disk and stream it back in chunks — keeps peak
        # memory flat even when the xlsx itself is tens of megabytes.
        tmp = tempfile.NamedTemporaryFile(prefix="report_", suffix=".xlsx", delete=False)
        tmp_path = tmp.name
        tmp.close()
        wb.save(tmp_path)

        def _iter_and_cleanup(path):
            try:
                with open(path, "rb") as fh:
                    while True:
                        chunk = fh.read(64 * 1024)
                        if not chunk:
                            break
                        yield chunk
            finally:
                try:
                    os.unlink(path)
                except OSError:
                    pass

        response = FileResponse(
            _iter_and_cleanup(tmp_path),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Length"] = str(os.path.getsize(tmp_path))
        return response

    except Exception as e:
        import traceback
        print("🔥 EXCEL EXPORT ERROR:")
        print(traceback.format_exc())
        from django.http import JsonResponse
        return JsonResponse({"error": str(e)}, status=500)